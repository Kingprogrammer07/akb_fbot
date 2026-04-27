"""
Bulk cargo photo report sender handler.

This module handles sending cargo photo reports to multiple clients
for a specific flight with progress tracking and error handling.
"""


from __future__ import annotations

import asyncio
import contextlib
import html as html_module
import json
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from time import monotonic
from typing import TYPE_CHECKING

from aiogram import Bot, F, Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputMediaPhoto,
    KeyboardButton,
    ReplyKeyboardMarkup,
    WebAppInfo,
)
from aiogram.exceptions import TelegramForbiddenError
from sqlalchemy.ext.asyncio import AsyncSession

from aiogram.fsm.state import State, StatesGroup as _StatesGroupAlias  # noqa: F401  (re-exported below)
from aiogram.types import Message

from src.bot.filters.is_admin import IsAdmin
from src.bot.handlers.admin._partner_alias_review import (
    FlightAliasReview,
    build_review,
    render_review_keyboard,
    render_review_text,
)
from src.bot.utils.flights_cache import get_flights_cache
from src.bot.utils.google_sheets_checker import GoogleSheetsChecker
from src.bot.utils.safe_sender import safe_execute, safe_send_message, safe_send_photo
from src.config import config
from src.infrastructure.database.client import DatabaseClient
from src.infrastructure.database.dao.client import ClientDAO
from src.infrastructure.database.dao.expected_cargo import ExpectedFlightCargoDAO
from src.infrastructure.database.dao.flight_cargo import FlightCargoDAO
from src.infrastructure.database.dao.partner import PartnerDAO
from src.infrastructure.database.dao.partner_payment_method import (
    PartnerPaymentMethodDAO,
)
from src.infrastructure.database.dao.partner_static_data import (
    PartnerStaticDataDAO,
)
from src.infrastructure.database.dao.static_data import StaticDataDAO
from src.infrastructure.services.flight_mask import (
    FlightMaskConflictError,
    FlightMaskError,
    FlightMaskService,
)
from src.infrastructure.services.partner_resolver import (
    PartnerNotFoundError,
    get_resolver,
)
from src.infrastructure.tools.s3_manager import s3_manager

if TYPE_CHECKING:
    from collections.abc import Sequence

logger = logging.getLogger(__name__)

router = Router(name="bulk_cargo_sender")

# Constants
MAX_PHOTOS_PER_MESSAGE = 10
PROGRESS_UPDATE_INTERVAL = 5  # percent
DEFAULT_USD_TO_UZS_RATE = 12_000
CAPTION_MAX_LENGTH = 1024


class SendStatus(str, Enum):
    """Status of a single send operation."""

    SUCCESS = "success"
    FAILED = "failed"
    BLOCKED = "blocked"


class BulkSendStates(StatesGroup):
    """FSM states for bulk cargo sending workflow."""

    selecting_flight = State()
    reviewing_aliases = State()
    editing_partner_alias = State()
    confirming_send = State()
    sending_in_progress = State()


@dataclass
class ErrorRecord:
    """Record of a failed send operation."""

    client_id: str
    flight_name: str
    error_reason: str
    timestamp: str = field(
        default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )


@dataclass
class SendStats:
    """Statistics for bulk send operation."""

    total: int = 0
    processed: int = 0
    sent: int = 0
    failed: int = 0
    blocked: int = 0
    start_time: float = field(default_factory=monotonic)
    errors: list[ErrorRecord] = field(default_factory=list)

    @property
    def progress_percent(self) -> int:
        return 0 if self.total == 0 else int((self.processed / self.total) * 100)

    @property
    def elapsed_time(self) -> float:
        return monotonic() - self.start_time

    @property
    def estimated_remaining(self) -> float:
        if self.processed == 0:
            return 0.0
        avg_time = self.elapsed_time / self.processed
        return (self.total - self.processed) * avg_time

    def should_update_progress(self) -> bool:
        """Check if progress should be updated based on interval."""
        if self.total == 0:
            return False
        interval = max(1, self.total // (100 // PROGRESS_UPDATE_INTERVAL))
        return self.processed % interval == 0 or self.processed == self.total

    def add_error(self, client_id: str, flight_name: str, reason: str) -> None:
        """Add an error record."""
        self.errors.append(
            ErrorRecord(
                client_id=client_id,
                flight_name=flight_name,
                error_reason=reason,
            )
        )


@dataclass
class BulkSendTask:
    """Represents an active bulk send task."""

    task: asyncio.Task
    cancelled: bool = False

    def cancel(self) -> None:
        self.cancelled = True


# Active bulk send tasks registry
_active_tasks: dict[str, BulkSendTask] = {}


def _generate_task_id(user_id: int) -> str:
    """Generate unique task ID."""
    return f"{user_id}_{int(monotonic() * 1000)}"


def _format_duration(seconds: float) -> str:
    """Format seconds to human-readable duration."""
    if seconds < 60:
        return f"{int(seconds)}s"
    if seconds < 3600:
        return f"{int(seconds // 60)}m {int(seconds % 60)}s"
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    return f"{hours}h {minutes}m"


def _build_keyboard(*buttons: tuple[str, str]) -> InlineKeyboardMarkup:
    """Build inline keyboard from button tuples (text, callback_data)."""
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=text, callback_data=data)]
            for text, data in buttons
        ]
    )


def _build_row_keyboard(buttons: list[tuple[str, str]]) -> InlineKeyboardMarkup:
    """Build inline keyboard with buttons in a single row."""
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text=text, callback_data=data)
                for text, data in buttons
            ]
        ]
    )


@dataclass
class CargoItemData:
    """Data for a single cargo item in the report."""

    weight: float
    price_usd: float
    price_uzs: float
    category: str = "Yuk"  # Default category name


@dataclass
class CargoReportData:
    """Data required to build and send a cargo report.

    ``flight_name`` is always the **real** flight name (used for DB writes,
    success/fail logs, mark_as_sent, transactions).  ``display_flight_name``
    is what is rendered in the message body delivered to end users — it is
    the partner-specific mask resolved via ``FlightMaskService``.  When no
    mask is configured both fields are equal.
    """

    CAPTION_LIMIT: int = field(default=1000, init=False, repr=False)

    client_id: str
    telegram_id: (
        int | None
    )  # Optional - only needed for sending to user, not for logging
    # True when a Client row was found in the DB for this client_id; False when
    # the code exists only in flight_cargos but has no matching Client record.
    client_in_db: bool
    track_codes: list[str]  # ALL track codes from Google Sheets (shown at the top)
    items: list[CargoItemData]  # List of individual cargo items
    total_weight: float
    extra_charge: float
    payment_card_number: str | None
    payment_card_holder: str | None
    foto_hisobot: str
    photo_file_ids: list[str]
    cargo_ids: list[int]
    flight_name: str
    display_flight_name: str = ""
    payment_links: list[tuple[str, str]] = field(default_factory=list)
    """Optional list of ``(label, url)`` tuples for online-payment providers
    (Click, Payme, …) supplied via ``partner_payment_methods``.  Rendered
    after the card block and **only** when the partner has at least one
    active link configured."""

    def __post_init__(self) -> None:
        if not self.display_flight_name:
            # Default: render real name when no mask was provided.
            self.display_flight_name = self.flight_name

    @property
    def total_price_uzs(self) -> float:
        """Calculate total price of all items in UZS."""
        return sum(item.price_uzs for item in self.items)

    @property
    def total_payment(self) -> float:
        """Total payment including extra charge."""
        return self.total_price_uzs + self.extra_charge

    def build_message(self) -> str:
        """Build the FULL message text (uses the *display* flight name)."""
        # Escape all user-supplied strings so stray '<', '>', '&' characters
        # in track codes, client IDs, or names do not break Telegram's HTML parser.
        safe_flight = html_module.escape(self.display_flight_name)
        safe_client = html_module.escape(self.client_id)

        # Build track codes section (shown at the top, ALL codes)
        track_info = ""
        if self.track_codes:
            track_info = (
                "<b>Trek kodlari:</b>\n"
                + "\n".join(
                    f"• <code>{html_module.escape(code)}</code>"
                    for code in self.track_codes
                )
                + "\n\n"
            )

        items_text = "".join(
            f"📦 <b>{html_module.escape(item.category)} #{idx}</b>\n⚖️ Vazn: {item.weight:.2f} kg\n💰 Narx: {item.price_usd:,.2f} $ ({item.price_uzs:,.0f} so'm)\n\n"
            for idx, item in enumerate(self.items, 1)
        )
        card_info = ""
        if self.payment_card_number:
            card_info = (
                f"💳 Karta raqami: <code>{html_module.escape(self.payment_card_number)}</code>\n"
                f"👤 Karta egasi: {html_module.escape(self.payment_card_holder or '')}\n"
            )

        link_info = ""
        if self.payment_links:
            link_lines = "\n".join(
                f"🌐 <a href=\"{html_module.escape(url, quote=True)}\">"
                f"{html_module.escape(label)}</a>"
                for label, url in self.payment_links
            )
            link_info = f"{link_lines}\n"

        message = (
            f"Assalomu aleykum. Yuqorida {safe_flight} - reysimizda kelgan "
            f"tovarlaringiz trek kodlari va foto hisoboti tashlandi.\n\n"
            f"{track_info}"
            f"{items_text}"
            f"<b>Mijoz kodi:</b> {safe_client}\n"
            f"<b>Jami vazn:</b> {self.total_weight:.2f} kg\n"
            f"<b>JAMI TO'LOV:</b> {self.total_payment:,.0f} so'm\n\n"
            f"{card_info}"
            f"{link_info}"
        )

        if self.foto_hisobot:
            message += f"\n{html_module.escape(self.foto_hisobot)}"

        return message


class ChannelLogger:
    """Handles logging to success/fail Telegram channels."""

    def __init__(self, bot: Bot):
        self.bot = bot
        self.success_channel = config.telegram.FOTO_HISOBOT_SUCCESS_CHANNEL_ID
        self.fail_channel = config.telegram.FOTO_HISOBOT_FAIL_CHANNEL_ID

    async def log_success(
        self,
        flight_name: str,
        client_id: str,
        telegram_id: int,
        message_text: str,
        photo_file_ids: list[str],
    ) -> None:
        """Log successful send to success channel."""
        header = self._build_header(
            status="✅ <b>Muvaffaqiyatli yuborildi</b>",
            flight_name=flight_name,
            client_id=client_id,
            telegram_id=telegram_id,
            message_text=message_text,
        )
        await self._send_log(self.success_channel, header, photo_file_ids)

    async def log_failure(
        self,
        flight_name: str,
        client_id: str,
        error: str,
        message_text: str = "",
        photo_file_ids: list[str] | None = None,
        reply_markup: InlineKeyboardMarkup | None = None,
    ) -> None:
        """Log failed send to fail channel."""
        header = (
            f"❌ <b>Yuborishda xatolik</b>\n\n"
            f"✈️ Reys: {html_module.escape(flight_name)}\n"
            f"👤 Mijoz: {html_module.escape(client_id)}\n"
            f"⚠️ Sabab: {html_module.escape(error)}\n"
            f"🕐 Vaqt: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"{message_text}"
        )
        await self._send_log(
            self.fail_channel, header, photo_file_ids or [], reply_markup
        )

    def _build_header(
        self,
        status: str,
        flight_name: str,
        client_id: str,
        telegram_id: int,
        message_text: str,
    ) -> str:
        return (
            f"{status}\n\n"
            f"✈️ Reys: {html_module.escape(flight_name)}\n"
            f"👤 Mijoz: {html_module.escape(client_id)}\n"
            f"🆔 Telegram ID: <code>{telegram_id}</code>\n"
            f"🕐 Vaqt: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"{message_text}"
        )

    async def _send_log(
        self,
        channel_id: int,
        text: str,
        photo_file_ids: list[str],
        reply_markup: InlineKeyboardMarkup | None = None,
    ) -> None:
        """
        Send log message.
        CRITICAL FIX: If reply_markup is present (or text is too long),
        we MUST send the text as a separate message because MediaGroups cannot have buttons.
        """
        try:
            is_too_long = len(text) > CAPTION_MAX_LENGTH
            has_button = reply_markup is not None

            # Force separate text if it's too long OR if we need to show a button
            force_separate_message = is_too_long or has_button

            if force_separate_message:
                caption = "📸 <b>Foto hisobot</b> (Batafsil ma'lumot pastda 👇)"
                send_full_text_after = True
            else:
                caption = text
                send_full_text_after = False

            last_msg_id = None

            # 1. Send Media
            if photo_file_ids:
                if len(photo_file_ids) == 1:
                    msg = await safe_send_photo(
                        self.bot,
                        chat_id=channel_id,
                        photo=photo_file_ids[0],
                        caption=caption,
                        parse_mode="HTML",
                    )
                    if msg:
                        last_msg_id = msg.message_id
                else:
                    media = [
                        InputMediaPhoto(
                            media=file_id,
                            caption=caption if idx == 0 else None,
                            parse_mode="HTML" if idx == 0 else None,
                        )
                        for idx, file_id in enumerate(
                            photo_file_ids[:MAX_PHOTOS_PER_MESSAGE]
                        )
                    ]
                    msgs = await safe_execute(
                        self.bot.send_media_group, chat_id=channel_id, media=media
                    )
                    if msgs and isinstance(msgs, list):
                        last_msg_id = msgs[-1].message_id

            # 2. Send Text (as Reply if media exists)
            if not photo_file_ids or send_full_text_after:
                await safe_send_message(
                    self.bot,
                    chat_id=channel_id,
                    text=text,
                    reply_to_message_id=last_msg_id,
                    parse_mode="HTML",
                    reply_markup=reply_markup,
                )

        except Exception as e:
            logger.warning("Failed to log to channel %s: %s", channel_id, e)
            # Fallback: strip all HTML tags so partial/broken markup can't cause
            # a second parse error, then send as plain text.
            try:
                plain_text = re.sub(r"<[^>]+>", "", text)[:500]
                await safe_send_message(
                    self.bot,
                    chat_id=channel_id,
                    text=plain_text,
                    reply_markup=reply_markup,
                )
            except Exception:
                logger.exception("Fallback log also failed")


class ProgressReporter:
    """Handles progress updates during bulk send."""

    def __init__(self, bot: Bot, chat_id: int, message_id: int, flight_name: str):
        self.bot = bot
        self.chat_id = chat_id
        self.message_id = message_id
        self.flight_name = flight_name

    async def update(self, stats: SendStats, task_id: str) -> None:
        """Update progress message."""
        text = (
            f"🚀 <b>Yuborish davom etmoqda...</b>\n\n"
            f"✈️ Reys: {self.flight_name}\n"
            f"👥 Jami: {stats.total}\n"
            f"📤 Yuborilmoqda: {stats.processed}/{stats.total} ({stats.progress_percent}%)\n"
            f"✅ Muvaffaqiyatli: {stats.sent}\n"
            f"❌ Xato: {stats.failed}\n"
            f"🚫 Bloklangan: {stats.blocked}\n\n"
            f"⏱ Qolgan vaqt: {_format_duration(stats.estimated_remaining)}"
        )

        keyboard = _build_keyboard(("⏸ Bekor qilish", f"bulk_cancel_task:{task_id}"))

        try:
            await safe_execute(
                self.bot.edit_message_text,
                chat_id=self.chat_id,
                message_id=self.message_id,
                text=text,
                parse_mode="HTML",
                reply_markup=keyboard,
            )
        except Exception as e:
            logger.warning("Failed to update progress: %s", e)
            await self._try_create_new_message(stats)

    async def finalize(self, stats: SendStats) -> None:
        """Send final summary message."""
        text = (
            f"✅ <b>Yuborish yakunlandi!</b>\n\n"
            f"✈️ Reys: {self.flight_name}\n"
            f"👥 Jami mijozlar: {stats.total}\n"
            f"✅ Muvaffaqiyatli: {stats.sent}\n"
            f"❌ Xato: {stats.failed}\n"
            f"🚫 Bloklangan: {stats.blocked}\n\n"
            f"⏱ Jami vaqt: {_format_duration(stats.elapsed_time)}"
        )

        try:
            await safe_execute(
                self.bot.edit_message_text,
                chat_id=self.chat_id,
                message_id=self.message_id,
                text=text,
                parse_mode="HTML",
            )
        except Exception as e:
            logger.warning("Failed to finalize progress: %s", e)

    async def _try_create_new_message(self, stats: SendStats) -> None:
        """Create new progress message if edit fails."""
        try:
            text = (
                f"🚀 <b>Yuborish davom etmoqda...</b>\n\n"
                f"✈️ Reys: {self.flight_name}\n"
                f"📤 {stats.processed}/{stats.total} ({stats.progress_percent}%)\n"
                f"✅ {stats.sent} | ❌ {stats.failed} | 🚫 {stats.blocked}"
            )
            new_msg = await safe_send_message(
                self.bot, chat_id=self.chat_id, text=text, parse_mode="HTML"
            )
            if new_msg:
                self.message_id = new_msg.message_id
        except Exception:
            logger.exception("Failed to create new progress message")


class ErrorReportGenerator:
    """Generates Excel report for failed sends."""

    def __init__(self, bot: Bot, admin_chat_id: int):
        self.bot = bot
        self.admin_chat_id = admin_chat_id

    async def generate_and_send(
        self, errors: list[ErrorRecord], flight_name: str
    ) -> None:
        """Generate Excel report and send to admin."""
        if not errors:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Xatoliklar"

            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill("solid", fgColor="C0392B")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Headers
            headers = ["#", "Mijoz ID", "Reys", "Xatolik sababi", "Vaqt"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Data rows
            for row_idx, error in enumerate(errors, 2):
                ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
                ws.cell(
                    row=row_idx, column=2, value=error.client_id
                ).border = thin_border
                ws.cell(
                    row=row_idx, column=3, value=error.flight_name
                ).border = thin_border
                ws.cell(
                    row=row_idx, column=4, value=error.error_reason
                ).border = thin_border
                ws.cell(
                    row=row_idx, column=5, value=error.timestamp
                ).border = thin_border

            # Column widths
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 50
            ws.column_dimensions["E"].width = 20

            # Summary row
            summary_row = len(errors) + 3
            ws.cell(row=summary_row, column=1, value="Jami xatoliklar:").font = Font(
                bold=True
            )
            ws.cell(row=summary_row, column=2, value=len(errors)).font = Font(bold=True)

            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"/tmp/xatoliklar_{flight_name}_{timestamp}.xlsx"
            wb.save(filename)

            # Send to admin
            from aiogram.types import FSInputFile

            doc = FSInputFile(filename, filename=f"xatoliklar_{flight_name}.xlsx")
            await safe_execute(
                self.bot.send_document,
                chat_id=self.admin_chat_id,
                document=doc,
                caption=(
                    f"📊 <b>Xatoliklar hisoboti</b>\n\n"
                    f"✈️ Reys: {flight_name}\n"
                    f"❌ Jami xatoliklar: {len(errors)}\n\n"
                    f"Batafsil ma'lumot Excel faylda."
                ),
                parse_mode="HTML",
            )

            # Cleanup
            import os

            os.remove(filename)

        except Exception as e:
            logger.exception("Failed to generate error report")
            # Fallback: send text summary
            await self._send_text_fallback(errors, flight_name)

    async def _send_text_fallback(
        self, errors: list[ErrorRecord], flight_name: str
    ) -> None:
        """Send text summary if Excel generation fails."""
        error_lines = [f"• {e.client_id}: {e.error_reason[:50]}" for e in errors[:20]]
        text = (
            f"📊 <b>Xatoliklar hisoboti</b>\n\n"
            f"✈️ Reys: {flight_name}\n"
            f"❌ Jami: {len(errors)}\n\n" + "\n".join(error_lines)
        )
        if len(errors) > 20:
            text += f"\n\n... va yana {len(errors) - 20} ta"

        await safe_send_message(
            self.bot, chat_id=self.admin_chat_id, text=text, parse_mode="HTML"
        )


class BulkCargoSender:
    """Main class for handling bulk cargo sending operations."""

    def __init__(
        self,
        bot: Bot,
        flight_name: str,
        clients_data: dict[str, list[int]],
        admin_chat_id: int,
        task_id: str,
    ):
        self.bot = bot
        self.flight_name = flight_name
        self.clients_data = clients_data
        self.admin_chat_id = admin_chat_id
        self.task_id = task_id
        self.stats = SendStats(total=len(clients_data))
        self.channel_logger = ChannelLogger(bot)
        self.error_reporter = ErrorReportGenerator(bot, admin_chat_id)
        self.progress_reporter: ProgressReporter | None = None
        self.db_client: DatabaseClient | None = None

        # Initialize Google Sheets checker for track codes
        self.sheets_checker = GoogleSheetsChecker(
            spreadsheet_id=config.google_sheets.SHEETS_ID,
            api_key=config.google_sheets.API_KEY,
            last_n_sheets=5,
        )

    async def initialize(self, progress_message_id: int) -> bool:
        """Initialize sender with progress reporter."""
        self.progress_reporter = ProgressReporter(
            self.bot, self.admin_chat_id, progress_message_id, self.flight_name
        )

        # Resolve real worksheet name from Google Sheets
        self._resolved_sheet_name: str | None = await self._resolve_sheet_name()
        if self._resolved_sheet_name:
            logger.info(
                "Resolved flight %s to worksheet %s",
                self.flight_name,
                self._resolved_sheet_name,
            )
        else:
            logger.warning(
                "Could not resolve worksheet for flight %s", self.flight_name
            )

        return True

    async def _resolve_sheet_name(self) -> str | None:
        """
        Resolve short flight code to full Google Sheets worksheet name.

        e.g., "M150" -> "M150-2025" or "M150-01-12"
        """
        try:
            # Get recent flight sheet names from Google Sheets
            sheet_names = await self.sheets_checker.get_flight_sheet_names(last_n=10)

            # Find worksheet that starts with our flight code
            flight_code_upper = self.flight_name.strip().upper()
            for sheet_name in sheet_names:
                if sheet_name.strip().upper().startswith(flight_code_upper):
                    return sheet_name

            # If no match, return the original flight_name as fallback
            logger.warning(
                "No matching worksheet found for %s in sheets: %s",
                self.flight_name,
                sheet_names,
            )
            return None
        except Exception as e:
            logger.exception("Failed to resolve sheet name for %s", self.flight_name)
            return None

    async def run(self) -> SendStats:
        """Execute bulk send operation."""
        async with DatabaseClient(config.database.database_url) as db_client:
            self.db_client = db_client
            try:
                static_data = await self._load_static_data()
                if not static_data:
                    return self.stats

                foto_hisobot, extra_charge = static_data

                for client_id, cargo_ids in self.clients_data.items():
                    if self._is_cancelled():
                        break

                    self.stats.processed += 1
                    await self._process_client(
                        client_id, cargo_ids, foto_hisobot, extra_charge
                    )

                    if self.stats.should_update_progress() and self.progress_reporter:
                        await self.progress_reporter.update(self.stats, self.task_id)

                if self.progress_reporter:
                    await self.progress_reporter.finalize(self.stats)

                # Generate error report if there are errors
                if self.stats.errors:
                    await self.error_reporter.generate_and_send(
                        self.stats.errors, self.flight_name
                    )

            finally:
                await self._cleanup()

        return self.stats

    def _is_cancelled(self) -> bool:
        """Check if task was cancelled."""
        task = _active_tasks.get(self.task_id)
        return task.cancelled if task else False

    async def _load_static_data(self) -> tuple[str, float] | None:
        """Load static data from database."""
        try:
            async with self.db_client.session_factory() as session:
                static_data = await StaticDataDAO.get_first(session)
                foto_hisobot = static_data.foto_hisobot if static_data else ""
                extra_charge = float(static_data.extra_charge) if static_data else 0.0
                return foto_hisobot, extra_charge
        except Exception as e:
            logger.exception("Failed to load static data")
            await safe_send_message(
                self.bot,
                chat_id=self.admin_chat_id,
                text=f"❌ Xatolik: Static data yuklanmadi: {e!s}",
            )
            return None

    async def _process_client(
        self,
        client_id: str,
        cargo_ids: list[int],
        foto_hisobot: str,
        extra_charge: float,
    ) -> None:
        """
        Process a single client's cargo report.

        CRITICAL: Only mark as sent if SUCCESS.
        """
        report_data: CargoReportData | None = None
        async with self.db_client.session_factory() as session:
            try:
                # STEP 1: Build full report data (does NOT require telegram_id)
                report_data = await self._build_report_data(
                    session, client_id, cargo_ids, foto_hisobot, extra_charge
                )

                if not report_data:
                    # No data to report
                    # We might want to mark as sent to avoid loops, OR keep pending
                    # For now, if NO DATA, we mark as sent to skip
                    await FlightCargoDAO.mark_as_sent(session, cargo_ids)
                    await session.commit()
                    return

                # STEP 2: Resolve the owning partner from the prefix.  Non-DM
                # partners (including ``GGX`` for the AKB Xorazm filiali) are
                # forwarded to ``partner.group_chat_id``; DM partners (AKB)
                # fall through to the direct-message flow below.
                try:
                    partner = await get_resolver().resolve_by_client_code(
                        session, client_id
                    )
                except PartnerNotFoundError as exc:
                    error_reason = f"Partner not registered: {exc!s}"
                    self.stats.failed += 1
                    self.stats.add_error(client_id, self.flight_name, error_reason)
                    await self.channel_logger.log_failure(
                        self.flight_name,
                        client_id,
                        error_reason,
                        message_text=report_data.build_message(),
                        photo_file_ids=report_data.photo_file_ids,
                    )
                    return

                if not partner.is_dm_partner:
                    await self._send_to_partner_group(
                        session, partner, client_id, report_data
                    )
                    return

                # STEP 3: AKB DM flow — requires telegram_id
                if not report_data.telegram_id:
                    # Cannot send to user -> Log failure with Manual Confirm button
                    error_reason = "Client not found or no telegram_id"
                    self.stats.failed += 1
                    self.stats.add_error(client_id, self.flight_name, error_reason)

                    # Create Manual Confirm Button
                    cb_data = f"manual_sent:{self.flight_name}:{client_id}"
                    keyboard = InlineKeyboardMarkup(
                        inline_keyboard=[
                            [
                                InlineKeyboardButton(
                                    text="✅ Qo'lda yuborildi", callback_data=cb_data
                                )
                            ]
                        ]
                    )

                    await self.channel_logger.log_failure(
                        self.flight_name,
                        client_id,
                        error_reason,
                        message_text=report_data.build_message(),
                        photo_file_ids=report_data.photo_file_ids,
                        reply_markup=keyboard,
                    )
                    return

                # STEP 3: Send to user (normal flow)
                result = await self._send_report(report_data)
                await self._handle_send_result(session, result, report_data)

            except Exception as e:
                # Catch-all: Do NOT mark as sent on crash, allow retry
                logger.exception("Error processing client %s", client_id)
                self.stats.failed += 1
                await self._log_exception(client_id, e, report_data)

    async def _build_report_data(
        self,
        session: AsyncSession,
        client_id: str,
        cargo_ids: list[int],
        foto_hisobot: str,
        extra_charge: float,
    ) -> CargoReportData | None:
        """
        Build cargo report data for a client.

        Track codes are ALWAYS fetched from Google Sheets (not DB).
        """
        # Get client (may be None — that's OK for building report)
        client = await ClientDAO.get_by_client_code(session, client_id)
        client_in_db: bool = client is not None
        telegram_id = client.telegram_id if client else None

        # Get cargos (required for report)
        cargos = await self._fetch_cargos(session, cargo_ids)
        if not cargos:
            return None

        # Build items list from cargos (NO track codes in items)
        items: list[CargoItemData] = []
        for cargo in cargos:
            weight = float(cargo.weight_kg or 0)
            price_per_kg_usd = float(cargo.price_per_kg or 0)
            price_per_kg_uzs = await self._convert_to_uzs(session, price_per_kg_usd)

            # Total price for this item
            total_price_usd = price_per_kg_usd * weight
            total_price_uzs = price_per_kg_uzs * weight

            items.append(
                CargoItemData(
                    weight=weight,
                    price_usd=total_price_usd,
                    price_uzs=total_price_uzs,
                    category="Yuk",
                )
            )

        # Collect track codes from BOTH sources simultaneously and merge them.
        # DB is authoritative; Google Sheets fills in any codes not yet imported
        # into expected_flight_cargos (legacy flights, missing entries, etc.).
        # Deduplication is by normalised upper-case value so the same physical
        # parcel is never listed twice even when both sources know about it.
        db_track_codes: list[str] = (
            await ExpectedFlightCargoDAO.get_track_codes_by_flight_and_client(
                session, self.flight_name, client_id
            )
        )

        sheets_track_codes: list[str] = []
        if self._resolved_sheet_name:
            try:
                sheets_track_codes = (
                    await self.sheets_checker.get_track_codes_by_flight_and_client(
                        flight_name=self._resolved_sheet_name,
                        client_code=client_id,
                    )
                )
            except Exception as e:
                await session.rollback()
                logger.warning(
                    "Failed to fetch track codes from Sheets for client %s: %s",
                    client_id,
                    e,
                )

        # Merge: DB codes first (preserve their order), then any Sheets codes
        # that are not already present in the DB set.
        seen_upper: set[str] = {c.upper() for c in db_track_codes}
        track_codes: list[str] = list(db_track_codes)
        for code in sheets_track_codes:
            if code.upper() not in seen_upper:
                track_codes.append(code)
                seen_upper.add(code.upper())

        # Resolve the partner-specific payment methods (cards + links).  The
        # resolver is called once more here even though _process_client also
        # uses it; PartnerResolver's internal cache makes this a near-free
        # dict lookup.  Falls back to the global ``payment_cards`` pool when
        # the partner has no card configured yet — preserves current
        # behaviour for any partner whose admin has not migrated their cards
        # to the new ``partner_payment_methods`` table.
        partner_payment_card = None
        partner_payment_links: list[tuple[str, str]] = []
        try:
            _partner_for_payment = await get_resolver().resolve_by_client_code(
                session, client_id
            )
        except PartnerNotFoundError:
            _partner_for_payment = None

        if _partner_for_payment is not None:
            partner_payment_card = (
                await PartnerPaymentMethodDAO.get_random_active_card(
                    session, _partner_for_payment.id
                )
            )
            partner_payment_links = [
                (link.link_label or "", link.link_url or "")
                for link in await PartnerPaymentMethodDAO.list_active_links(
                    session, _partner_for_payment.id
                )
            ]

        if partner_payment_card is not None:
            payment_card_number = partner_payment_card.card_number
            payment_card_holder = partner_payment_card.card_holder
        else:
            # Partner has no active card configured.  The message renders
            # without a card block; admins are expected to add at least one
            # method via the admin panel before the next send.
            payment_card_number = None
            payment_card_holder = None

        # Per-partner foto_hisobot override.  Empty string in the per-partner
        # row means "no override"; fall back to the singleton ``StaticData``
        # value (already resolved by the caller and passed in as foto_hisobot).
        partner_foto_hisobot = foto_hisobot
        if _partner_for_payment is not None:
            psd = await PartnerStaticDataDAO.get_for_partner(
                session, _partner_for_payment.id
            )
            if psd and psd.foto_hisobot:
                partner_foto_hisobot = psd.foto_hisobot

        # Calculate totals
        total_weight = sum(item.weight for item in items)

        # Extract photos and resolve S3 keys to presigned URLs
        raw_photo_ids = self._extract_photos(cargos)
        if not raw_photo_ids:
            return None
        photo_file_ids = await self._resolve_photo_references(raw_photo_ids)
        if not photo_file_ids:
            return None

        # Resolve the partner-specific mask for this client so the message
        # body shows the alias rather than the real flight code.  The DAO
        # returns ``None`` when no alias has been configured (which the
        # admin-side review flow normally guarantees), in which case the
        # CargoReportData default falls back to the real flight name.
        display_flight = self.flight_name
        try:
            partner = await get_resolver().resolve_by_client_code(
                session, client_id
            )
        except PartnerNotFoundError:
            partner = None
        if partner is not None:
            mask = await FlightMaskService.real_to_mask(
                session, partner.id, self.flight_name
            )
            if mask:
                display_flight = mask

        return CargoReportData(
            client_id=client_id,
            telegram_id=telegram_id,
            client_in_db=client_in_db,
            track_codes=track_codes,  # ALL track codes (shown at top)
            items=items,  # Individual cargo items (no track codes)
            total_weight=total_weight,
            extra_charge=extra_charge,
            payment_card_number=payment_card_number,
            payment_card_holder=payment_card_holder,
            foto_hisobot=partner_foto_hisobot,
            photo_file_ids=photo_file_ids,
            cargo_ids=cargo_ids,
            flight_name=self.flight_name,
            display_flight_name=display_flight,
            payment_links=partner_payment_links,
        )

    async def _fetch_cargos(self, session: AsyncSession, cargo_ids: list[int]) -> list:
        """Fetch cargo records by IDs."""
        cargos = []
        for cargo_id in cargo_ids:
            cargo = await FlightCargoDAO.get_by_id(session, cargo_id)
            if cargo:
                cargos.append(cargo)
        return cargos

    async def _convert_to_uzs(self, session: AsyncSession, usd_amount: float) -> float:
        """Convert USD to UZS with fallback and db check."""
        try:
            from src.bot.utils.currency_converter import currency_converter

            rate = await currency_converter.get_rate_async(session, "USD", "UZS")
            return usd_amount * rate
        except Exception:
            await session.rollback()
            logger.warning("Currency conversion failed, using fallback rate")
            return usd_amount * DEFAULT_USD_TO_UZS_RATE

    def _extract_photos(self, cargos: Sequence) -> list[str]:
        """Extract photo file IDs from cargos."""
        photos = []
        for cargo in cargos:
            try:
                cargo_photos = json.loads(cargo.photo_file_ids or "[]")
                photos.extend(cargo_photos[:MAX_PHOTOS_PER_MESSAGE])
            except (json.JSONDecodeError, TypeError):
                continue
        return photos[:MAX_PHOTOS_PER_MESSAGE]

    async def _resolve_photo_references(self, items: list[str]) -> list[str]:
        """Resolve a mixed list of photo references for use with Aiogram.

        S3 keys (contain ``/``) are converted to temporary presigned URLs.
        Legacy Telegram file_ids (no ``/``) are returned as-is.

        Args:
            items: List of S3 keys or Telegram file_ids.

        Returns:
            List of resolved references (URLs or raw file_ids) ready for Aiogram.
        """
        resolved: list[str] = []
        for item in items:
            if "/" in item:
                try:
                    url = await s3_manager.generate_presigned_url(item, expires_in=3600)
                    resolved.append(url)
                except Exception as exc:
                    logger.error(
                        "Failed to generate presigned URL for S3 key %s: %s", item, exc
                    )
            else:
                resolved.append(item)
        return resolved

    async def _send_report(
        self,
        report_data: CargoReportData,
        override_chat_id: int | None = None,
        include_payment_keyboard: bool = True,
    ) -> dict:
        """
        Send cargo report: Photos first, then Text (replying to photos).
        Uses safe_sender utilities to handle flood limits and network errors.

        Args:
            report_data: Full cargo report payload.
            override_chat_id: When set, send to this chat instead of
                report_data.telegram_id (used for group sends, e.g. Xorazm filiali).
            include_payment_keyboard: Whether to attach the "💳 To'lov qilish"
                ReplyKeyboard.  Set to False when sending to groups/channels.
        """
        chat_id: int = (
            override_chat_id
            if override_chat_id is not None
            else report_data.telegram_id
        )
        last_msg_id = None

        # 1. Send Photos (if any)
        if report_data.photo_file_ids:
            try:
                # We use a short caption for photos to avoid limit issues
                caption = "📸 Foto hisobot"

                if len(report_data.photo_file_ids) == 1:
                    msg = await safe_send_photo(
                        self.bot,
                        chat_id=chat_id,
                        photo=report_data.photo_file_ids[0],
                        caption=caption,
                    )
                    if msg:
                        last_msg_id = msg.message_id
                else:
                    # InputMediaPhoto for album
                    media = [
                        InputMediaPhoto(media=pid, caption=caption if i == 0 else None)
                        for i, pid in enumerate(
                            report_data.photo_file_ids[:MAX_PHOTOS_PER_MESSAGE]
                        )
                    ]
                    msgs = await safe_execute(
                        self.bot.send_media_group, chat_id=chat_id, media=media
                    )
                    # Reply to the last message in the group
                    if msgs:
                        last_msg_id = msgs[-1].message_id

            except TelegramForbiddenError:
                return {
                    "success": False,
                    "status": "blocked",
                    "error": "Bot blocked by user",
                }
            except Exception as e:
                logger.warning(
                    f"Photo send failed for client {report_data.client_id}: {e}"
                )
                # Don't return, try sending text anyway as it contains the critical info

        # 2. Send Text (Reply)
        try:
            full_text = report_data.build_message()

            payment_keyboard = (
                ReplyKeyboardMarkup(
                    keyboard=[[KeyboardButton(text="💳 To'lov qilish")]],
                    resize_keyboard=True,
                    one_time_keyboard=False,
                )
                if include_payment_keyboard
                else None
            )

            await safe_send_message(
                self.bot,
                chat_id=chat_id,
                text=full_text,
                reply_to_message_id=last_msg_id,
                parse_mode="HTML",
                reply_markup=payment_keyboard,
            )
            return {"success": True, "status": "sent"}

        except TelegramForbiddenError:
            return {
                "success": False,
                "status": "blocked",
                "error": "Bot blocked by user",
            }
        except Exception as e:
            return {"success": False, "status": "error", "error": str(e)}

    async def _handle_send_result(
        self,
        session: AsyncSession,
        result: dict,
        report_data: CargoReportData,
    ) -> None:
        """Handle the result of a send operation."""
        log_message = report_data.build_message()

        if result.get("success"):
            # 1. Mark as sent in DB (ONLY HERE - on SUCCESS)
            await self._mark_as_sent(session, report_data)

            # 2. Update Stats
            self.stats.sent += 1

            # 3. Log to Success Channel
            await self.channel_logger.log_success(
                self.flight_name,
                report_data.client_id,
                report_data.telegram_id,
                log_message,
                report_data.photo_file_ids,
            )
        else:
            # Handle Fail/Block - Do NOT update DB
            status = result.get("status", "")
            error_reason = result.get("error", "Unknown error")

            if status == "blocked":
                self.stats.blocked += 1
                self.stats.add_error(
                    report_data.client_id, self.flight_name, f"Blocked: {error_reason}"
                )
            else:
                self.stats.failed += 1
                self.stats.add_error(
                    report_data.client_id, self.flight_name, error_reason
                )

            # Create Manual Confirm Button for Fail Channel
            cb_data = f"manual_sent:{self.flight_name}:{report_data.client_id}"
            keyboard = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="✅ Qo'lda yuborildi", callback_data=cb_data
                        )
                    ]
                ]
            )

            # Send to Fail Channel so Admin knows!
            await self.channel_logger.log_failure(
                self.flight_name,
                report_data.client_id,
                error_reason,
                log_message,
                report_data.photo_file_ids,
                reply_markup=keyboard,
            )

    async def _mark_as_sent(
        self, session: AsyncSession, report_data: CargoReportData
    ) -> None:
        """Mark cargos as sent and create a debt transaction."""
        await FlightCargoDAO.mark_as_sent(session, report_data.cargo_ids)

        # Accrual Accounting: Create debt transaction for the client immediately
        # Use ClientTransactionDAO to log the debt for this flight
        from src.infrastructure.database.dao.client_transaction import (
            ClientTransactionDAO,
        )

        # Dublikat tekshiruvi: client mavjud bo'lsa barcha aktiv kodlarini
        # (extra_code / client_code / legacy_code) hisobga olib qidiramiz.
        # Shunda reys bo'yicha 1 ta user uchun faqat 1 ta qator saqlanadi.
        from src.infrastructure.services.client import ClientService

        client = await ClientService().get_client_by_code(
            report_data.client_id, session
        )
        lookup_codes = (
            client.active_codes if client else [report_data.client_id]
        )
        existing_tx = await ClientTransactionDAO.get_by_client_code_flight(
            session,
            lookup_codes,
            report_data.flight_name,
        )

        if not existing_tx:
            await ClientTransactionDAO.create(
                session,
                {
                    "telegram_id": report_data.telegram_id or 0,
                    "client_code": report_data.client_id,
                    "qator_raqami": 0,
                    "reys": report_data.flight_name,
                    "summa": 0,
                    "vazn": str(round(report_data.total_weight, 2)),
                    "payment_type": "online",
                    "payment_status": "pending",
                    "paid_amount": 0,
                    "total_amount": report_data.total_payment,
                    "remaining_amount": report_data.total_payment,
                    "payment_balance_difference": -report_data.total_payment,
                    "is_taken_away": False,
                },
            )

        await session.commit()

    async def _send_to_partner_group(
        self,
        session: AsyncSession,
        partner,
        client_id: str,
        report_data: CargoReportData,
    ) -> None:
        """Forward a cargo report to a non-DM partner's Telegram group.

        ``partner.is_dm_partner`` is False here.  When ``group_chat_id`` is
        not configured the send fails loudly so the admin notices and adds
        the missing ID via the partner admin panel.
        """
        group_id = partner.group_chat_id
        if not group_id:
            error_reason = (
                f"Partner {partner.code!r} has no group_chat_id configured"
            )
            self.stats.failed += 1
            self.stats.add_error(client_id, self.flight_name, error_reason)
            await self.channel_logger.log_failure(
                self.flight_name,
                client_id,
                error_reason,
                message_text=report_data.build_message(),
                photo_file_ids=report_data.photo_file_ids,
            )
            return

        result = await self._send_report(
            report_data,
            override_chat_id=group_id,
            include_payment_keyboard=False,
        )

        if result.get("success"):
            await self._mark_as_sent(session, report_data)
            self.stats.sent += 1
            logger.info(
                "_send_to_partner_group: partner=%r client=%r flight=%r → group %d",
                partner.code,
                client_id,
                self.flight_name,
                group_id,
            )
            await self.channel_logger.log_success(
                self.flight_name,
                client_id,
                group_id,
                report_data.build_message(),
                report_data.photo_file_ids,
            )
        else:
            error_reason = result.get("error", "Guruhga yuborishda noma'lum xato")
            self.stats.failed += 1
            self.stats.add_error(
                client_id,
                self.flight_name,
                f"{partner.code}→guruh xato: {error_reason}",
            )
            await self.channel_logger.log_failure(
                self.flight_name,
                client_id,
                f"{partner.display_name} guruhiga yuborishda xato: {error_reason}",
                message_text=report_data.build_message(),
                photo_file_ids=report_data.photo_file_ids,
            )

    async def _log_exception(
        self,
        client_id: str,
        error: Exception,
        report_data: CargoReportData | None = None,
    ) -> None:
        """Log exception to fail channel and admin."""
        error_reason = f"Exception: {error!s}"
        self.stats.add_error(client_id, self.flight_name, error_reason)

        # Build message text from report_data if available
        message_text = ""
        photo_file_ids: list[str] = []
        if report_data:
            message_text = report_data.build_message()
            photo_file_ids = report_data.photo_file_ids

        # Create Manual Confirm Button
        cb_data = f"manual_sent:{self.flight_name}:{client_id}"
        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="✅ Qo'lda yuborildi", callback_data=cb_data
                    )
                ]
            ]
        )

        await self.channel_logger.log_failure(
            self.flight_name,
            client_id,
            error_reason,
            message_text=message_text,
            photo_file_ids=photo_file_ids,
            reply_markup=keyboard,
        )
        with contextlib.suppress(Exception):
            await safe_send_message(
                self.bot,
                chat_id=self.admin_chat_id,
                text=f"⚠️ Xatolik ({client_id}): {str(error)[:200]}",
            )

    async def _cleanup(self) -> None:
        """Cleanup resources."""
        _active_tasks.pop(self.task_id, None)


# Router handlers


@router.callback_query(F.data == "start_bulk_send", IsAdmin())
async def start_bulk_send(
    callback: CallbackQuery, state: FSMContext, bot: Bot, session: AsyncSession
):
    """Start bulk cargo sending - show flight selection.

    Flight list is built from two sources merged without duplicates:
    1. Google Sheets recent flights (existing cache).
    2. Distinct flight names from the expected_flight_cargos DB table.
    """
    await callback.answer()

    # Source 1: Google Sheets flights (cached)
    sheets_flights: list[str] = await get_flights_cache().get_flights()

    # Source 2: Last 5 expected-cargo DB flights (most recently updated first)
    expected_flight_stats = await ExpectedFlightCargoDAO.get_distinct_flights(
        session, limit=5
    )
    db_flight_names: list[str] = [fs.flight_name for fs in expected_flight_stats]

    if not sheets_flights and not db_flight_names:
        await callback.message.answer("❌ Hech qanday reys topilmadi!")
        return

    # Build two separate sections — no deduplication across sources so admin
    # can clearly see which flights exist in each system.
    rows: list[list[InlineKeyboardButton]] = []
    if sheets_flights:
        rows.append([InlineKeyboardButton(text="━━ 📊 Google Sheets ━━", callback_data="noop")])
        for flight in sheets_flights:
            rows.append([InlineKeyboardButton(
                text=f"✈️ {flight}",
                callback_data=f"bulk_select_flight:{flight}",
            )])
    if db_flight_names:
        rows.append([InlineKeyboardButton(text="━━ 🗄 Expected Cargo DB ━━", callback_data="noop")])
        for flight in db_flight_names:
            rows.append([InlineKeyboardButton(
                text=f"✈️ {flight}",
                callback_data=f"bulk_select_flight:{flight}",
            )])
    rows.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="bulk_cancel")])
    keyboard = InlineKeyboardMarkup(inline_keyboard=rows)

    await callback.message.delete()

    await safe_send_message(
        bot,
        chat_id=callback.from_user.id,
        text=(
            "📋 <b>Qaysi reys uchun ma'lumot yubormoqchisiz?</b>\n\n"
            "📊 Google Sheets — foto yuklangan reyslar\n"
            "🗄 Expected Cargo DB — oldindan kiritilgan reyslar\n\n"
            "Kerakli reysni tanlang:"
        ),
        reply_markup=keyboard,
        parse_mode="HTML",
    )

    await state.set_state(BulkSendStates.selecting_flight)


@router.callback_query(F.data.startswith("bulk_select_flight:"), IsAdmin())
async def select_flight(
    callback: CallbackQuery,
    state: FSMContext,
    session: AsyncSession,
    bot: Bot,
):
    """Flight selected - show confirmation."""
    await callback.answer()
    flight_name = callback.data.split(":", 1)[1]

    cargos = await FlightCargoDAO.get_unsent_by_flight(session, flight_name)

    if not cargos:
        await safe_send_message(
            bot,
            chat_id=callback.from_user.id,
            text=f"❌ {flight_name} reysi uchun yuborilmagan yuklar topilmadi!",
        )
        await state.clear()
        return

    # Group by client_id
    clients_data: dict[str, list[int]] = {}
    for cargo in cargos:
        clients_data.setdefault(cargo.client_id, []).append(cargo.id)

    total_clients = len(clients_data)
    total_cargos = len(cargos)

    await state.update_data(
        flight_name=flight_name,
        clients_data=clients_data,
        total_clients=total_clients,
        total_cargos=total_cargos,
    )

    # Build the per-partner mask review (auto-generates default aliases for
    # any partner that does not yet have one for this flight).
    review = await build_review(
        session, real_flight_name=flight_name, client_codes=clients_data.keys()
    )
    await session.commit()

    await callback.message.delete()

    await _render_review_screen(bot, callback.from_user.id, flight_name, review)
    await state.set_state(BulkSendStates.reviewing_aliases)


# ──────────────────────────────────────────────────────────────
# Per-partner alias review handlers
# ──────────────────────────────────────────────────────────────


async def _render_review_screen(
    bot: Bot,
    chat_id: int,
    flight_name: str,
    review: FlightAliasReview,
    extra_note: str | None = None,
) -> None:
    """Render the alias review screen (used both initially and after edits)."""
    text = render_review_text(review)
    if extra_note:
        text = f"{extra_note}\n\n{text}"
    keyboard = render_review_keyboard(
        review,
        edit_callback_prefix="bulk_alias_edit",
        proceed_callback="bulk_alias_proceed",
        cancel_callback="bulk_cancel",
    )
    await safe_send_message(
        bot,
        chat_id=chat_id,
        text=text,
        reply_markup=keyboard,
        parse_mode="HTML",
    )


@router.callback_query(
    F.data.startswith("bulk_alias_edit:"),
    BulkSendStates.reviewing_aliases,
    IsAdmin(),
)
async def alias_edit_prompt(
    callback: CallbackQuery,
    state: FSMContext,
    session: AsyncSession,
    bot: Bot,
):
    """Admin chose a partner row → ask for the new mask."""
    await callback.answer()
    try:
        partner_id = int(callback.data.split(":", 1)[1])
    except (ValueError, IndexError):
        await callback.answer("❌ Noto'g'ri partner ID", show_alert=True)
        return

    partner = await PartnerDAO.get_by_id(session, partner_id)
    if partner is None:
        await callback.answer("❌ Partner topilmadi", show_alert=True)
        return

    data = await state.get_data()
    flight_name = data.get("flight_name")
    if not flight_name:
        await callback.answer("❌ Reys topilmadi, qaytadan boshlang", show_alert=True)
        await state.clear()
        return

    current_mask = await FlightMaskService.real_to_mask(
        session, partner.id, flight_name
    )

    await state.update_data(editing_partner_id=partner.id)
    await state.set_state(BulkSendStates.editing_partner_alias)

    await callback.message.delete()
    await safe_send_message(
        bot,
        chat_id=callback.from_user.id,
        text=(
            f"✏️ <b>{html_module.escape(partner.display_name)}</b> "
            f"({html_module.escape(partner.code)}) uchun yangi maska kiriting.\n\n"
            f"<b>Haqiqiy reys:</b> <code>{html_module.escape(flight_name)}</code>\n"
            f"<b>Hozirgi maska:</b> "
            f"<code>{html_module.escape(current_mask or '— belgilanmagan —')}</code>\n\n"
            f"Yangi maskani matn shaklida yuboring (1–100 belgi)."
        ),
        parse_mode="HTML",
    )


@router.message(BulkSendStates.editing_partner_alias, IsAdmin(), F.text)
async def alias_edit_save(
    message: Message,
    state: FSMContext,
    session: AsyncSession,
    bot: Bot,
):
    """Receive the new mask, validate, persist, then redraw the review."""
    raw = (message.text or "").strip()
    if not raw or len(raw) > 100:
        await message.answer("❌ Maska 1–100 belgi bo'lishi kerak. Qaytadan kiriting.")
        return

    data = await state.get_data()
    flight_name = data.get("flight_name")
    partner_id = data.get("editing_partner_id")
    clients_data: dict[str, list[int]] = data.get("clients_data") or {}

    if not flight_name or not partner_id:
        await message.answer("❌ Sessiya yo'qoldi. Qaytadan boshlang.")
        await state.clear()
        return

    try:
        await FlightMaskService.set_mask(
            session,
            partner_id=partner_id,
            real_flight_name=flight_name,
            new_mask=raw,
        )
        await session.commit()
    except FlightMaskConflictError as exc:
        await message.answer(
            f"❌ Bu maska allaqachon band: <code>{html_module.escape(str(exc))}</code>\n"
            f"Boshqa nom kiriting.",
            parse_mode="HTML",
        )
        return
    except FlightMaskError as exc:
        await message.answer(f"❌ Xatolik: {exc!s}")
        return

    review = await build_review(
        session, real_flight_name=flight_name, client_codes=clients_data.keys()
    )
    await state.set_state(BulkSendStates.reviewing_aliases)
    await state.update_data(editing_partner_id=None)
    await _render_review_screen(
        bot,
        message.from_user.id,
        flight_name,
        review,
        extra_note="✅ Maska yangilandi.",
    )


@router.callback_query(
    F.data == "bulk_alias_proceed",
    BulkSendStates.reviewing_aliases,
    IsAdmin(),
)
async def alias_proceed(
    callback: CallbackQuery,
    state: FSMContext,
    bot: Bot,
):
    """Admin confirmed all masks → show the final send confirmation screen."""
    await callback.answer()

    data = await state.get_data()
    flight_name = data.get("flight_name")
    total_clients = data.get("total_clients", 0)
    total_cargos = data.get("total_cargos", 0)
    if not flight_name:
        await callback.answer("❌ Reys topilmadi, qaytadan boshlang", show_alert=True)
        await state.clear()
        return

    keyboard = _build_row_keyboard(
        [
            ("✅ Yuborish", "bulk_confirm_send"),
            ("❌ Bekor qilish", "bulk_cancel"),
        ]
    )

    await callback.message.delete()
    await safe_send_message(
        bot,
        chat_id=callback.from_user.id,
        text=(
            f"📊 <b>Yuborish tasdigi</b>\n\n"
            f"✈️ Reys: <b>{html_module.escape(flight_name)}</b>\n"
            f"👥 Mijozlar soni: <b>{total_clients}</b>\n"
            f"📦 Jami yuklar: <b>{total_cargos}</b>\n\n"
            f"Barcha mijozlarga foto hisobot yuborilsinmi?"
        ),
        reply_markup=keyboard,
        parse_mode="HTML",
    )

    await state.set_state(BulkSendStates.confirming_send)


@router.callback_query(F.data == "bulk_confirm_send", IsAdmin())
async def confirm_send(callback: CallbackQuery, state: FSMContext, bot: Bot):
    """Confirmed - start bulk sending.

    Dispatches to the ostatka pipeline when the flight name starts with
    ``A-``; otherwise uses the regular ``BulkCargoSender``.  Both senders
    share identical external signatures so the handler stays thin.
    """
    await callback.answer()

    data = await state.get_data()
    flight_name = data["flight_name"]
    clients_data = data["clients_data"]
    total_clients = data["total_clients"]

    # Local import keeps the ostatka module optional and avoids a cycle —
    # ostatka_sender itself imports helpers defined above in this file.
    from src.bot.handlers.admin.ostatka_sender import (
        OstatkaBulkSender,
        is_ostatka_flight,
    )

    is_ostatka = is_ostatka_flight(flight_name)
    header = "♻️ <b>Ostatka yuborish boshlandi...</b>" if is_ostatka else "🚀 <b>Yuborish boshlandi...</b>"

    # Create progress message
    progress_msg = await callback.message.answer(
        f"{header}\n\n"
        f"✈️ Reys: {flight_name}\n"
        f"👥 Jami: {total_clients}\n"
        f"📤 Yuborilmoqda: 0/{total_clients} (0%)\n"
        f"✅ Muvaffaqiyatli: 0\n"
        f"❌ Xato: 0\n"
        f"🚫 Bloklangan: 0\n\n"
        f"⏱ Taxminiy vaqt: Hisoblanyapti...",
        parse_mode="HTML",
    )

    task_id = _generate_task_id(callback.from_user.id)

    await progress_msg.edit_reply_markup(
        reply_markup=_build_keyboard(
            ("⏸ Bekor qilish", f"bulk_cancel_task:{progress_msg.message_id}")
        )
    )

    # Create sender and start task
    sender_cls = OstatkaBulkSender if is_ostatka else BulkCargoSender
    sender = sender_cls(
        bot=bot,
        flight_name=flight_name,
        clients_data=clients_data,
        admin_chat_id=callback.from_user.id,
        task_id=task_id,
    )
    await sender.initialize(progress_msg.message_id)

    task = asyncio.create_task(sender.run())
    _active_tasks[task_id] = BulkSendTask(task=task)

    await state.set_state(BulkSendStates.sending_in_progress)
    await state.update_data(task_id=task_id)


@router.callback_query(F.data.startswith("bulk_cancel_task:"), IsAdmin())
async def cancel_task(callback: CallbackQuery, state: FSMContext):
    """Cancel ongoing bulk send."""
    await callback.answer("Bekor qilinyapti...")

    data = await state.get_data()
    task_id = data.get("task_id")

    if task_id and task_id in _active_tasks:
        _active_tasks[task_id].cancel()

    await callback.message.delete()
    await state.clear()


@router.callback_query(F.data == "noop", IsAdmin())
async def noop_handler(callback: CallbackQuery):
    """No-op handler for section header buttons."""
    await callback.answer()


@router.callback_query(F.data == "bulk_cancel", IsAdmin())
async def cancel_bulk_send(callback: CallbackQuery, state: FSMContext):
    """Cancel bulk send process."""
    await callback.answer()
    await callback.message.answer("❌ Bekor qilindi")
    await state.clear()
    await callback.message.delete()


@router.callback_query(F.data.startswith("manual_sent:"), IsAdmin())
async def manual_sent_handler(
    callback: CallbackQuery,
    session: AsyncSession,
    bot: Bot,
):
    """
    Handle manual confirmation of sent cargo.

    Format: manual_sent:{flight_name}:{client_id}
    """
    try:
        _, flight_name, client_id = callback.data.split(":", 2)

        # 1. Update DB: Mark as sent
        # Get all unsent cargos for this client+flight
        cargos = await FlightCargoDAO.get_unsent_by_flight(session, flight_name)
        # Filter for specific client
        client_cargos = [c for c in cargos if c.client_id == client_id]
        if unsent_ids := [c.id for c in client_cargos if not c.is_sent]:
            await FlightCargoDAO.mark_as_sent(session, unsent_ids)
            
            # --- Qarz yozish logikasi ---
            from src.infrastructure.services.client import ClientService
            from src.infrastructure.database.dao.client_transaction import ClientTransactionDAO
            from src.bot.utils.currency_converter import currency_converter
            from src.infrastructure.database.dao.static_data import StaticDataDAO
            from src.bot.handlers.admin.bulk_cargo_sender import DEFAULT_USD_TO_UZS_RATE
            
            client = await ClientService().get_client_by_code(client_id, session)
            lookup_codes = client.active_codes if client else [client_id]
            telegram_id = client.telegram_id if client else 0
            
            # Kurs va qadoqlash narxini olish
            static_data = await StaticDataDAO.get_first(session)
            extra_charge = float(static_data.extra_charge) if static_data else 0.0
            try:
                rate = await currency_converter.get_rate_async(session, "USD", "UZS")
            except Exception:
                rate = DEFAULT_USD_TO_UZS_RATE

            total_weight = 0.0
            total_price_uzs = 0.0
            for c in client_cargos:
                if c.id in unsent_ids:
                    w = float(c.weight_kg or 0)
                    p_usd = float(c.price_per_kg or 0)
                    total_weight += w
                    total_price_uzs += (p_usd * rate) * w
            
            total_payment = total_price_uzs + extra_charge
            
            existing_tx = await ClientTransactionDAO.get_by_client_code_flight(
                session, lookup_codes, flight_name
            )
            
            if not existing_tx:
                await ClientTransactionDAO.create(
                    session,
                    {
                        "telegram_id": telegram_id,
                        "client_code": client_id,
                        "qator_raqami": 0,
                        "reys": flight_name,
                        "summa": 0,
                        "vazn": str(round(total_weight, 2)),
                        "payment_type": "online",
                        "payment_status": "pending",
                        "paid_amount": 0,
                        "total_amount": total_payment,
                        "remaining_amount": total_payment,
                        "payment_balance_difference": -total_payment,
                        "is_taken_away": False,
                    }
                )
            else:
                current_total = float(existing_tx.total_amount or existing_tx.summa or 0)
                existing_tx.total_amount = current_total + total_payment
                
                current_remaining = float(existing_tx.remaining_amount or 0)
                existing_tx.remaining_amount = current_remaining + total_payment
                
                current_diff = float(existing_tx.payment_balance_difference or 0)
                existing_tx.payment_balance_difference = current_diff - total_payment
                
                try:
                    current_vazn = float(existing_tx.vazn) if existing_tx.vazn else 0.0
                except ValueError:
                    current_vazn = 0.0
                existing_tx.vazn = str(round(current_vazn + total_weight, 2))
                
                if existing_tx.payment_status == "paid" and total_payment > 0:
                    existing_tx.payment_status = "partial"

            await session.commit()
            await callback.answer(
                f"✅ {len(unsent_ids)} ta yuk belgilandi va qarz yozildi!", show_alert=True
            )
        else:
            await callback.answer(
                "⚠️ Yuborilmagan yuklar topilmadi (allaqachon belgilangan bo'lishi mumkin)",
                show_alert=True,
            )

        # 2. Edit message to remove button and add "Manually Sent" tag
        original_text = (
                    callback.message.html_text or callback.message.caption
                ) or (callback.message.text or "")

        # Check if already marked (to avoid double edits)
        if "✅ <b>--- Admin tomonidan qo'lda yuborildi ---</b>" in original_text:
            return

        new_text = (
            original_text + "\n\n✅ <b>--- Admin tomonidan qo'lda yuborildi ---</b>"
        )

        # Edit based on message type
        try:
            if callback.message.photo:
                await bot.edit_message_caption(
                    chat_id=callback.message.chat.id,
                    message_id=callback.message.message_id,
                    caption=new_text,
                    parse_mode="HTML",
                    reply_markup=None,
                )
            else:
                await bot.edit_message_text(
                    chat_id=callback.message.chat.id,
                    message_id=callback.message.message_id,
                    text=new_text,
                    parse_mode="HTML",
                    reply_markup=None,
                )
        except Exception as e:
            await session.rollback()
            if "message is not modified" not in str(e):
                logger.warning("Failed to edit manual send message: %s", e)

    except Exception as e:
        await session.rollback()
        logger.exception("Manual sent handler error")
        await callback.answer("❌ Xatolik yuz berdi", show_alert=True)


# ──────────────────────────────────────────────────────────────
# Web Report Send Handlers
# ──────────────────────────────────────────────────────────────


@router.callback_query(F.data == "start_web_bulk_send", IsAdmin())
async def start_web_bulk_send(
    callback: CallbackQuery, state: FSMContext, bot: Bot, session: AsyncSession
):
    """Start web report send - show flight selection.

    Flight list is built from two sources merged without duplicates:
    1. Google Sheets recent flights (existing cache).
    2. Last 5 most-recently-updated distinct flights from expected_flight_cargos DB.
    """
    await callback.answer()

    # Source 1: Google Sheets flights (cached)
    sheets_flights: list[str] = await get_flights_cache().get_flights()

    # Source 2: Last 5 expected-cargo DB flights (most recently updated first)
    expected_flight_stats = await ExpectedFlightCargoDAO.get_distinct_flights(
        session, limit=5
    )
    db_flight_names: list[str] = [fs.flight_name for fs in expected_flight_stats]

    if not sheets_flights and not db_flight_names:
        await callback.message.answer("❌ Hech qanday reys topilmadi!")
        return

    # Build two separate sections — no deduplication across sources so admin
    # can clearly see which flights exist in each system.
    rows: list[list[InlineKeyboardButton]] = []
    if sheets_flights:
        rows.append([InlineKeyboardButton(text="━━ 📊 Google Sheets ━━", callback_data="noop")])
        for flight in sheets_flights:
            rows.append([InlineKeyboardButton(
                text=f"✈️ {flight}",
                callback_data=f"web_select_flight:{flight}",
            )])
    if db_flight_names:
        rows.append([InlineKeyboardButton(text="━━ 🗄 Expected Cargo DB ━━", callback_data="noop")])
        for flight in db_flight_names:
            rows.append([InlineKeyboardButton(
                text=f"✈️ {flight}",
                callback_data=f"web_select_flight:{flight}",
            )])
    rows.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="bulk_cancel")])
    keyboard = InlineKeyboardMarkup(inline_keyboard=rows)

    await callback.message.delete()

    await safe_send_message(
        bot,
        chat_id=callback.from_user.id,
        text=(
            "🌐 <b>Web hisobot yuborish</b>\n\n"
            "📊 Google Sheets — foto yuklangan reyslar\n"
            "🗄 Expected Cargo DB — oldindan kiritilgan reyslar\n\n"
            "Kerakli reysni tanlang:"
        ),
        reply_markup=keyboard,
        parse_mode="HTML",
    )


@router.callback_query(F.data.startswith("web_select_flight:"), IsAdmin())
async def web_select_flight(
    callback: CallbackQuery,
    state: FSMContext,
    session: AsyncSession,
    bot: Bot,
):
    """Web flight selected - show confirmation with stats."""
    await callback.answer()
    flight_name = callback.data.split(":", 1)[1]

    # Get cargos not yet sent to web
    cargos = await FlightCargoDAO.get_unsent_web_by_flight(session, flight_name)

    if not cargos:
        await safe_send_message(
            bot,
            chat_id=callback.from_user.id,
            text=f"❌ {flight_name} reysi uchun web'ga yuborilmagan yuklar topilmadi!",
        )
        await state.clear()
        return

    # Collect unique clients and cargo IDs
    unique_clients: set[str] = set()
    cargo_ids: list[int] = []
    for cargo in cargos:
        unique_clients.add(cargo.client_id)
        cargo_ids.append(cargo.id)

    total_clients = len(unique_clients)
    total_cargos = len(cargo_ids)

    # Store in FSM
    await state.update_data(
        web_flight_name=flight_name,
        web_cargo_ids=cargo_ids,
    )

    keyboard = _build_row_keyboard(
        [
            ("✅ Tasdiqlash", "web_confirm_send"),
            ("❌ Bekor qilish", "bulk_cancel"),
        ]
    )

    await callback.message.delete()

    await safe_send_message(
        bot,
        chat_id=callback.from_user.id,
        text=(
            f"📊 <b>Web hisobot yuborish</b>\n\n"
            f"✈️ Reys: <b>{flight_name}</b>\n"
            f"👥 Mijozlar: <b>{total_clients}</b>\n"
            f"📦 Yuklar: <b>{total_cargos}</b>\n\n"
            f"Ushbu ma'lumotlar web server uchun 'Yuborilgan' deb belgilansinmi?"
        ),
        reply_markup=keyboard,
        parse_mode="HTML",
    )


@router.callback_query(F.data == "web_confirm_send", IsAdmin())
async def web_confirm_send(
    callback: CallbackQuery,
    state: FSMContext,
    session: AsyncSession,
    bot: Bot,
):
    """Confirmed - mark cargos as sent to web."""
    await callback.answer()

    data = await state.get_data()
    flight_name = data.get("web_flight_name", "")
    cargo_ids = data.get("web_cargo_ids", [])

    if not cargo_ids:
        await callback.message.answer("❌ Yuklar topilmadi. Qaytadan urinib ko'ring.")
        await state.clear()
        return

    # Mark all as sent to web
    updated_count = await FlightCargoDAO.mark_as_sent_web(session, cargo_ids)
    await session.commit()

    # ── Notify clients ──────────────────────────────────────────
    # Get all cargos that were just marked (re-fetch to get client_ids)
    cargos = await FlightCargoDAO.get_by_ids(
        session, cargo_ids
    )  # sizda bor bo'lsa, yoki quyida loop

    # Group cargo_ids by client_id
    client_ids: set[str] = {cargo.client_id for cargo in cargos}

    notified = 0
    failed = 0
    for client_id in client_ids:
        client = await ClientDAO.get_by_client_code(session, client_id)
        if not client or not client.telegram_id:
            failed += 1
            continue

        try:
            sent = await safe_send_message(
                bot,
                chat_id=client.telegram_id,
                text=(
                    f"📦 Hurmatli mijoz!\n\n"
                    f"Sizga <b>{flight_name}</b> reysi foto hisoboti yuborildi. "
                    f"Hisobotni ko'rish uchun saytimizga kiring."
                ),
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [
                            InlineKeyboardButton(
                                text="📊 Hisobotni ko'rish",
                                web_app=WebAppInfo(
                                    url=config.telegram.webapp_login_url
                                ),
                            )
                        ]
                    ]
                ),
            )
            if sent:
                notified += 1
            else:
                # safe_send_message exhausted retries but did not raise
                failed += 1
        except Exception as exc:
            # TelegramForbiddenError (bot blocked), TelegramBadRequest (chat not
            # found / user deactivated), or network errors after max retries —
            # log and continue so the remaining clients still get notified.
            logger.warning(
                "web_confirm_send: could not notify client %s (telegram_id=%s): %s",
                client_id,
                client.telegram_id,
                exc,
            )
            failed += 1
    # ────────────────────────────────────────────────────────────

    await callback.message.delete()

    await safe_send_message(
        bot,
        chat_id=callback.from_user.id,
        text=(
            f"✅ <b>Web hisobot muvaffaqiyatli belgilandi!</b>\n\n"
            f"✈️ Reys: <b>{flight_name}</b>\n"
            f"📦 Belgilangan yuklar: <b>{updated_count}</b>\n"
            f"✅ Xabardor qilindi: <b>{notified}</b>\n"
            f"❌ Yuborilmadi (telegram yo'q): <b>{failed}</b>"
        ),
        parse_mode="HTML",
    )

    await state.clear()
