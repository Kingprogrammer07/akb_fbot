import io
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from src.api.schemas.statistics.cargo_stats import (
    CargoStatsResponse,
    CargoVolumeStats,
    CargoBottleneckStats,
    CargoSpeedStats,
    FlightVolumeItem,
    TrackSearchTrendItem,
)
from src.infrastructure.database.dao.statistics.cargo_stats import CargoStatsDAO


class CargoStatsService:
    def __init__(self, dao: CargoStatsDAO):
        self.dao = dao

    async def get_stats(
        self, start_date: Optional[datetime] = None, end_date: Optional[datetime] = None
    ) -> CargoStatsResponse:
        volume_data = await self.dao.get_volume_stats(start_date, end_date)
        bottleneck_data = await self.dao.get_bottleneck_stats(start_date, end_date)
        speed_data = await self.dao.get_speed_stats(start_date, end_date)
        top_flights_data = await self.dao.get_top_flights(
            start_date, end_date, limit=10
        )
        period_trends_data = await self.dao.get_period_trends(start_date, end_date)

        volume = CargoVolumeStats(**volume_data)
        bottlenecks = CargoBottleneckStats(**bottleneck_data)
        speed = CargoSpeedStats(**speed_data)
        top_flights = [FlightVolumeItem(**item) for item in top_flights_data]
        period_trends = [TrackSearchTrendItem(**item) for item in period_trends_data]

        return CargoStatsResponse(
            volume=volume,
            bottlenecks=bottlenecks,
            speed=speed,
            top_flights=top_flights,
            period_trends=period_trends,
        )

    async def export_stats_excel(
        self, start_date: Optional[datetime] = None, end_date: Optional[datetime] = None
    ) -> io.BytesIO:
        stats = await self.get_stats(start_date, end_date)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Yuk Statistikasi"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def write_header(sheet, row, col, text):
            cell = sheet.cell(row=row, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

        # 1. Asosiy hajmlar
        ws.cell(row=1, column=1, value="Asosiy yuk hajmlari").font = Font(
            bold=True, size=14
        )
        write_header(ws, 2, 1, "Ko'rsatkich")
        write_header(ws, 2, 2, "Qiymat")

        ws.cell(row=3, column=1, value="Jami kelgan yuklar soni")
        ws.cell(row=3, column=2, value=stats.volume.total_cargos)
        ws.cell(row=4, column=1, value="Jami kelgan yuklar vazni (kg)")
        ws.cell(row=4, column=2, value=float(stats.volume.total_weight_kg))
        ws.cell(row=5, column=1, value="Bitta mijozga o'rtacha vazn (kg)")
        ws.cell(row=5, column=2, value=float(stats.volume.avg_weight_per_client))
        ws.cell(row=6, column=1, value="Bitta paket (trek) uchun o'rtacha vazn (kg)")
        ws.cell(row=6, column=2, value=float(stats.volume.avg_weight_per_track))

        # 2. Muammoli nuqtalar
        ws.cell(row=8, column=1, value="Muammoli nuqtalar va holatlar").font = Font(
            bold=True, size=14
        )
        write_header(ws, 9, 1, "Ko'rsatkich")
        write_header(ws, 9, 2, "Soni")

        ws.cell(row=10, column=1, value="Xitoyda bor, UZ ga kiritilmagan")
        ws.cell(row=10, column=2, value=stats.bottlenecks.china_unaccounted)
        ws.cell(row=11, column=1, value="UZ da bor, to'lov kutilmoqda")
        ws.cell(row=11, column=2, value=stats.bottlenecks.uz_pending_payment)
        ws.cell(row=12, column=1, value="To'langan, olinmagan")
        ws.cell(row=12, column=2, value=stats.bottlenecks.uz_paid_not_taken)
        ws.cell(row=13, column=1, value="Mijoz olib ketgan")
        ws.cell(row=13, column=2, value=stats.bottlenecks.uz_taken_away)
        ws.cell(row=14, column=1, value="Pochtaga topshirilgan")
        ws.cell(row=14, column=2, value=stats.bottlenecks.post_approved)

        # 3. Aylanma tezlik
        ws.cell(row=16, column=1, value="Aylanma tezlik (kunlarda)").font = Font(
            bold=True, size=14
        )
        write_header(ws, 17, 1, "Ko'rsatkich")
        write_header(ws, 17, 2, "Kun")

        ws.cell(row=18, column=1, value="Xitoydan UZ ga o'rtacha kelish vaqti")
        ws.cell(row=18, column=2, value=float(stats.speed.china_to_uz_days))
        ws.cell(row=19, column=1, value="UZ da omborda turish vaqti")
        ws.cell(row=19, column=2, value=float(stats.speed.uz_warehouse_days))
        ws.cell(row=20, column=1, value="Umumiy aylanma tezlik")
        ws.cell(row=20, column=2, value=float(stats.speed.full_cycle_days))

        # 4. Top reyslar
        ws.cell(
            row=22, column=1, value="Eng katta hajmli reyslar (Top 10)"
        ).font = Font(bold=True, size=14)
        write_header(ws, 23, 1, "Reys nomi")
        write_header(ws, 23, 2, "Yuklar soni")
        write_header(ws, 23, 3, "Jami vazn (kg)")

        row_idx = 24
        for flight in stats.top_flights:
            ws.cell(row=row_idx, column=1, value=flight.flight_name)
            ws.cell(row=row_idx, column=2, value=flight.cargo_count)
            ws.cell(row=row_idx, column=3, value=float(flight.total_weight_kg))
            row_idx += 1

        # 5. Track kod qidiruv trendlari
        trend_row = row_idx + 2
        ws.cell(row=trend_row, column=1, value="Kunlik Track Kod Qidiruvlar").font = Font(
            bold=True, size=14
        )
        write_header(ws, trend_row + 1, 1, "Sana")
        write_header(ws, trend_row + 1, 2, "Qidiruvlar soni")

        row_idx = trend_row + 2
        for trend in stats.period_trends:
            ws.cell(row=row_idx, column=1, value=trend.period_name)
            ws.cell(row=row_idx, column=2, value=trend.search_count)
            row_idx += 1

        # Column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
