import contextlib
from datetime import date
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook

from src.infrastructure.database.dao.statistics.client_stats import ClientStatsDAO
from src.api.schemas.statistics.client_stats import (
    ClientStatsResponse,
    OverviewStats,
    RetentionStats,
    RegionDetail,
    DistrictDetail,
    DeliveryStatItem,
)


class ClientStatsService:
    """Service handling business logic for Client Statistics."""

    def __init__(self, session: AsyncSession):
        self.dao = ClientStatsDAO(session)

    async def get_stats(
        self, start_date: date | None, end_date: date | None
    ) -> ClientStatsResponse:
        """
        Fetches all statistics for clients and formats them into a Pydantic response.
        """
        overview_data = await self.dao.get_overview_and_retention(start_date, end_date)
        region_data = await self.dao.get_region_stats(start_date, end_date)
        delivery_data = await self.dao.get_delivery_stats(start_date, end_date)

        overview = OverviewStats(
            total_clients=overview_data["total_clients"],
            new_clients=overview_data["new_clients"],
            active_clients=overview_data["active_clients"],
            passive_clients=overview_data["passive_clients"],
            zombie_clients=overview_data["zombie_clients"],
            logged_in_clients=overview_data["logged_in_clients"],
        )

        retention = RetentionStats(
            repeat_clients=overview_data["repeat_clients"],
            one_time_clients=overview_data["one_time_clients"],
            most_frequent_clients=overview_data["most_frequent_clients"],
        )

        # region_data — DAO dan nested dict keladi:
        # { region_name: { code, count, revenue, paid, debt, districts: { d_name: {...} } } }
        regions: dict[str, RegionDetail] = {
            region_name: RegionDetail(
                code=r["code"],
                count=r["count"],
                revenue=r["revenue"],
                paid=r["paid"],
                debt=r["debt"],
                districts={
                    d_name: DistrictDetail(
                        code=d["code"],
                        count=d["count"],
                        revenue=d["revenue"],
                        paid=d["paid"],
                        debt=d["debt"],
                    )
                    for d_name, d in r["districts"].items()
                },
            )
            for region_name, r in region_data.items()
        }

        delivery_methods = [
            DeliveryStatItem(method=d["method"], count=d["count"])
            for d in delivery_data
        ]

        return ClientStatsResponse(
            overview=overview,
            retention=retention,
            regions=regions,
            delivery_methods=delivery_methods,
        )

    @staticmethod
    def _build_excel(rows: list[dict], columns: list[tuple[str, str]]) -> BytesIO:
        """Generic helper: builds a simple openpyxl workbook from rows + column spec."""
        wb = Workbook()
        ws = wb.active

        # Header row
        for col_idx, (_, header) in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Data rows
        for row_num, row in enumerate(rows, start=2):
            for col_idx, (key, _) in enumerate(columns, start=1):
                val = row.get(key)
                # Strip timezone info for Excel compatibility
                if hasattr(val, "replace") and hasattr(val, "tzinfo") and val.tzinfo:
                    val = val.replace(tzinfo=None)
                ws.cell(row=row_num, column=col_idx, value=val)

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    async def get_zombie_excel(
        self, start_date: date | None, end_date: date | None
    ) -> BytesIO:
        rows = await self.dao.get_zombie_clients(start_date, end_date)
        columns = [
            ("active_codes",  "Kodlar (extra | client | legacy)"),
            ("phone",         "Telefon"),
            ("created_at",    "Ro'yxatdan o'tgan sana"),
            ("updated_at",    "Yangilangan sana"),
            ("last_seen_at",  "Oxirgi bot faolligi"),
        ]
        return self._build_excel(rows, columns)

    async def get_passive_excel(
        self, start_date: date | None, end_date: date | None
    ) -> BytesIO:
        rows = await self.dao.get_passive_clients(start_date, end_date)
        columns = [
            ("active_codes",    "Kodlar (extra | client | legacy)"),
            ("phone",           "Telefon"),
            ("created_at",      "Ro'yxatdan o'tgan sana"),
            ("updated_at",      "Yangilangan sana"),
            ("last_seen_at",    "Oxirgi bot faolligi"),
            ("last_cargo_date", "Oxirgi yuk sanasi"),
            ("last_flight_name","Oxirgi reys"),
            ("last_weight_kg",  "Oxirgi yuk vazni (kg)"),
        ]
        return self._build_excel(rows, columns)

    async def get_frequent_excel(self, min_flights: int = 5) -> BytesIO:
        rows = await self.dao.get_frequent_clients(min_flights)
        columns = [
            ("active_codes",   "Kodlar (extra | client | legacy)"),
            ("phone",          "Telefon"),
            ("flight_count",   "Reyslar soni"),
            ("total_cargos",   "Jami yuklar"),
            ("total_weight_kg","Jami vazn (kg)"),
            ("created_at",     "Ro'yxatdan o'tgan sana"),
            ("updated_at",     "Yangilangan sana"),
            ("last_seen_at",   "Oxirgi bot faolligi"),
        ]
        return self._build_excel(rows, columns)

    async def get_stats_excel(
        self, start_date: date | None, end_date: date | None
    ) -> BytesIO:
        """
        Generates an Excel file containing all the client statistics.
        """
        stats = await self.get_stats(start_date, end_date)

        wb = Workbook()
        ws = wb.active
        ws.title = "Mijozlar Statistikasi"

        # Overview
        ws.append(["Mijozlar Statistikasi (Umumiy)"])
        ws.append(["Ko'rsatkich", "Soni"])
        ws.append(["Jami mijozlar", stats.overview.total_clients])
        ws.append(["Yangi ro'yxatdan o'tganlar", stats.overview.new_clients])
        ws.append(["Aktiv mijozlar (oxirgi 45 kun)", stats.overview.active_clients])
        ws.append(["Passiv mijozlar (60 kundan oshgan)", stats.overview.passive_clients])
        ws.append(["Zombi mijozlar (hech qachon yuk olmagan)", stats.overview.zombie_clients])
        ws.append(["Hozirda tizimda (is_logged_in)", stats.overview.logged_in_clients])
        ws.append([])

        # Retention
        ws.append(["Mijozlarni ushlab qolish (Retention)"])
        ws.append(["Ko'rsatkich", "Soni"])
        ws.append(["Qayta buyurtma qilganlar (Sodiq)", stats.retention.repeat_clients])
        ws.append(["Bir marta ishlatib ketganlar", stats.retention.one_time_clients])
        ws.append(["Eng ko'p buyurtma beradiganlar (5+ reys)", stats.retention.most_frequent_clients])
        ws.append([])

        # Regions + districts
        ws.append(["Hududlar bo'yicha ko'rsatkichlar"])
        ws.append(["Viloyat / Shahar", "Tuman", "Mijozlar", "Hisoblangan (so'm)", "To'langan (so'm)", "Qarz (so'm)"])
        for r_name, r in stats.regions.items():
            ws.append([r_name, "— JAMI —", r.count, r.revenue, r.paid, r.debt])
            for d_name, d in r.districts.items():
                ws.append(["", d_name, d.count, d.revenue, d.paid, d.debt])
        ws.append([])

        # Delivery Methods
        ws.append(["Yetkazib berish usullari bo'yicha"])
        ws.append(["Usul", "Soni"])
        for d in stats.delivery_methods:
            ws.append([d.method, d.count])

        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                with contextlib.suppress(Exception):
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            ws.column_dimensions[column].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
