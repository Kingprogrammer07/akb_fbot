from datetime import datetime
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.api.schemas.statistics.financial_stats import (
    FinancialStatsResponse,
    PeriodicRevenue,
    PaymentMethodStat,
    RegionStat,
    TopClientStat,
    FlightCollectionStat,
)
from src.api.utils.constants import decode_region_key
from src.infrastructure.database.dao.statistics.financial_stats import FinancialStatsDAO


class FinancialStatsService:
    @staticmethod
    async def get_summary(
        session: AsyncSession,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        base_cost_usd: float = 8.0,
    ) -> FinancialStatsResponse:
        summary = await FinancialStatsDAO.get_financial_summary(
            session, start_date, end_date, base_cost_usd
        )
        periodic = await FinancialStatsDAO.get_periodic_revenue(
            session, start_date, end_date
        )
        methods = await FinancialStatsDAO.get_payment_methods(
            session, start_date, end_date
        )
        regions_raw = await FinancialStatsDAO.get_regions(session, start_date, end_date)
        top_clients = await FinancialStatsDAO.get_top_clients(
            session, start_date, end_date
        )
        flight_cols = await FinancialStatsDAO.get_flight_collections(
            session, start_date, end_date
        )

        decoded_regions = [
            RegionStat(
                region_code=r["region_code"],
                region_name=decode_region_key(r["region_code"]),
                revenue=float(r.get("revenue") or 0.0),
                paid=float(r.get("paid") or 0.0),
                debt=float(r.get("debt") or 0.0),
            )
            for r in regions_raw
        ]

        return FinancialStatsResponse(
            total_revenue=summary.get("total_revenue", 0.0),
            total_paid=summary.get("total_paid", 0.0),
            total_debt=summary.get("total_debt", 0.0),
            total_profitability=summary.get("total_profitability", 0.0),
            overdue_debt=summary.get("overdue_debt", 0.0),
            average_payment=summary.get("average_payment", 0.0),
            periodic_revenue=[PeriodicRevenue(**r) for r in periodic],
            payment_methods=[PaymentMethodStat(**r) for r in methods],
            regions=decoded_regions,
            top_clients=[TopClientStat(**r) for r in top_clients],
            flight_collections=[FlightCollectionStat(**r) for r in flight_cols],
        )

    @staticmethod
    async def generate_excel(
        session: AsyncSession,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        base_cost_usd: float = 8.0,
    ) -> BytesIO:
        stats = await FinancialStatsService.get_summary(
            session, start_date, end_date, base_cost_usd
        )
        regions_hier = await FinancialStatsDAO.get_regions_hierarchical(
            session, start_date, end_date
        )

        wb = Workbook()

        # ── Varaq 1: Umumiy hisobot ──────────────────────────────────────────
        ws = wb.active
        ws.title = "Umumiy Hisobot"
        _header(ws, ["Ko'rsatkich", "Qiymat (so'm)"])
        rows_summary = [
            ("Jami hisoblangan (Billed)", stats.total_revenue),
            ("Jami to'langan",            stats.total_paid),
            ("Jami qarzdorlik",           stats.total_debt),
            ("Muddati o'tgan qarz (>15 kun)", stats.overdue_debt),
            ("O'rtacha to'lov",           stats.average_payment),
            ("Sof foyda (taxminiy)",      stats.total_profitability),
        ]
        for label, value in rows_summary:
            ws.append([label, round(value, 2)])
        _autowidth(ws)

        # ── Varaq 2: Hududlar (viloyat + tumanlar) ───────────────────────────
        ws2 = wb.create_sheet("Hududlar")
        _header(ws2, ["Viloyat / Shahar", "Tuman", "Hisoblangan (so'm)", "To'langan (so'm)", "Qarz (so'm)"])
        for region_name, r in regions_hier.items():
            # Viloyat jami qatori
            ws2.append([
                region_name, "— JAMI —",
                round(r["revenue"], 2),
                round(r["paid"],    2),
                round(r["debt"],    2),
            ])
            # Har bir tuman
            for d_name, d in r["districts"].items():
                ws2.append([
                    "", d_name,
                    round(d["revenue"], 2),
                    round(d["paid"],    2),
                    round(d["debt"],    2),
                ])
        _autowidth(ws2)

        # ── Varaq 3: To'lov usullari ─────────────────────────────────────────
        ws3 = wb.create_sheet("To'lov usullari")
        _header(ws3, ["Usul", "Jami summa (so'm)", "Soni"])
        for m in stats.payment_methods:
            ws3.append([m.method, round(m.total_amount, 2), m.count])
        _autowidth(ws3)

        # ── Varaq 4: Reyslar yig'imi ─────────────────────────────────────────
        ws4 = wb.create_sheet("Reyslar")
        _header(ws4, ["Reys", "Hisoblangan (so'm)", "To'langan (so'm)", "Undirilish (%)"])
        for f in stats.flight_collections:
            ws4.append([
                f.flight_name,
                round(f.revenue, 2),
                round(f.paid,    2),
                round(f.collection_rate, 1),
            ])
        _autowidth(ws4)

        # ── Varaq 5: Top mijozlar ────────────────────────────────────────────
        ws5 = wb.create_sheet("Top Mijozlar")
        _header(ws5, ["Mijoz kodi", "Hisoblangan (so'm)", "To'langan (so'm)", "Qarz (so'm)"])
        for c in stats.top_clients:
            ws5.append([c.client_code, round(c.revenue, 2), round(c.paid, 2), round(c.debt, 2)])
        _autowidth(ws5)

        # ── Varaq 6: Oylik dinamika ──────────────────────────────────────────
        ws6 = wb.create_sheet("Oylik Dinamika")
        _header(ws6, ["Oy", "Hisoblangan (so'm)", "To'langan (so'm)", "Qarz (so'm)"])
        for p in stats.periodic_revenue:
            ws6.append([p.period, round(p.revenue, 2), round(p.paid, 2), round(p.debt, 2)])
        _autowidth(ws6)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# ── Yordamchi funksiyalar ────────────────────────────────────────────────────

def _header(ws, titles: list[str]) -> None:
    """Qalin sarlavha qatori yozadi."""
    ws.append(titles)
    header_row = ws[ws.max_row]
    for cell in header_row:
        cell.font = Font(bold=True)


def _autowidth(ws) -> None:
    """Ustun kengligini tarkibga qarab moslashtiradi."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
