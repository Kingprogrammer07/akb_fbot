"""Business logic service for the Expected Cargo feature.

This layer sits between the Router and the DAO.  It:
  • Orchestrates multi-step operations (e.g. track-code → client lookup).
  • Generates the Excel workbook for streaming export.
  • Translates DAO-level value objects into API response schemas.
  • Raises HTTPException so that routers stay thin.
"""
import io
import logging
import math
from datetime import datetime

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from src.api.schemas.expected_cargo import (
    BulkCreateExpectedCargoResponse,
    ClientStatItem,
    ClientSummaryItem,
    DeleteExpectedCargoResponse,
    ExpectedCargoItem,
    ExpectedCargoSummaryStats,
    FlightListItem,
    FlightListResponse,
    FlightStatItem,
    PaginatedClientStatsResponse,
    PaginatedClientSummaryResponse,
    PaginatedExpectedCargoResponse,
    PaginatedFlightStatsResponse,
    RenameClientCodeResponse,
    RenameFlightResponse,
    ReplaceTrackCodesResponse,
    ResolvedClientResponse,
)
from src.infrastructure.database.dao.cargo_item import CargoItemDAO
from src.infrastructure.database.dao.client import ClientDAO
from src.infrastructure.database.dao.expected_cargo import ExpectedFlightCargoDAO

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Excel styling constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_GROUP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

_EXCEL_HEADERS = [
    "T/R",
    "Reys nomi",
    "Mijoz kodi",
    "Trek kodi",
    "Yaratilgan",
    "Yangilangan",
]


class ExpectedCargoService:
    """Service class for all Expected Cargo business operations."""

    # -------------------------------------------------------------------
    # API 1 — Bulk create
    # -------------------------------------------------------------------

    @staticmethod
    async def bulk_create(
        session: AsyncSession,
        flight_name: str,
        client_code: str,
        track_codes: list[str],
    ) -> BulkCreateExpectedCargoResponse:
        """
        Insert multiple track codes for a flight + client pair.

        Duplicate codes are skipped gracefully and reported in the response
        rather than causing a 409 error, because partial success is the norm
        during manual data entry.

        Args:
            session:     Active DB session (caller must commit).
            flight_name: Target flight name.
            client_code: Target client code.
            track_codes: List of tracking codes to insert.

        Returns:
            BulkCreateExpectedCargoResponse with counts and any duplicates.
        """
        result = await ExpectedFlightCargoDAO.bulk_create(
            session=session,
            flight_name=flight_name,
            client_code=client_code,
            track_codes=track_codes,
        )
        await session.commit()

        return BulkCreateExpectedCargoResponse(
            created_count=result.created_count,
            duplicate_track_codes=result.duplicate_track_codes,
        )

    # -------------------------------------------------------------------
    # API 2 — Paginated search
    # -------------------------------------------------------------------

    @staticmethod
    async def search(
        session: AsyncSession,
        page: int,
        size: int,
        flight_name: str | None,
        client_code: str | None,
        track_code: str | None,
    ) -> PaginatedExpectedCargoResponse:
        """
        Return a paginated, optionally-filtered list of expected cargo records.

        Args:
            session:     Active DB session.
            page:        1-based page number.
            size:        Records per page.
            flight_name: Optional flight filter.
            client_code: Optional client filter.
            track_code:  Optional partial track code filter.

        Returns:
            PaginatedExpectedCargoResponse.
        """
        records, total = await ExpectedFlightCargoDAO.paginated_search(
            session=session,
            page=page,
            size=size,
            flight_name=flight_name,
            client_code=client_code,
            track_code=track_code,
        )
        total_pages = math.ceil(total / size) if total else 0

        return PaginatedExpectedCargoResponse(
            items=[ExpectedCargoItem.model_validate(r) for r in records],
            total=total,
            page=page,
            size=size,
            total_pages=total_pages,
        )

    # -------------------------------------------------------------------
    # API 3 — Replace-all
    # -------------------------------------------------------------------

    @staticmethod
    async def replace_track_codes(
        session: AsyncSession,
        flight_name: str,
        client_code: str,
        new_track_codes: list[str],
    ) -> ReplaceTrackCodesResponse:
        """
        Atomically replace all track codes for a flight + client.

        Args:
            session:        Active DB session (caller must commit).
            flight_name:    Target flight.
            client_code:    Target client.
            new_track_codes: Full replacement list (may be empty to clear).

        Returns:
            ReplaceTrackCodesResponse with deleted and created counts.
        """
        result = await ExpectedFlightCargoDAO.replace_client_track_codes(
            session=session,
            flight_name=flight_name,
            client_code=client_code,
            new_track_codes=new_track_codes,
        )
        await session.commit()

        return ReplaceTrackCodesResponse(
            deleted_count=result.deleted_count,
            created_count=result.created_count,
        )

    # -------------------------------------------------------------------
    # API 4 — Rename flight
    # -------------------------------------------------------------------

    @staticmethod
    async def rename_flight(
        session: AsyncSession,
        old_flight_name: str,
        new_flight_name: str,
    ) -> RenameFlightResponse:
        """
        Bulk-rename a flight across all its expected cargo records.

        Raises 404 if no records are found for old_flight_name (prevents
        silent no-ops that could indicate a typo in the flight name).

        Args:
            session:         Active DB session.
            old_flight_name: Current flight name.
            new_flight_name: Replacement flight name.

        Returns:
            RenameFlightResponse with the updated row count.

        Raises:
            HTTPException 404: If old_flight_name matches no records.
        """
        updated_count = await ExpectedFlightCargoDAO.rename_flight(
            session=session,
            old_flight_name=old_flight_name,
            new_flight_name=new_flight_name,
        )

        if updated_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"'{old_flight_name}' nomli reys topilmadi.",
            )

        await session.commit()

        return RenameFlightResponse(
            updated_count=updated_count,
            old_flight_name=old_flight_name,
            new_flight_name=new_flight_name,
        )

    # -------------------------------------------------------------------
    # API 4b — Rename client code within a flight
    # -------------------------------------------------------------------

    @staticmethod
    async def rename_client_code(
        session: AsyncSession,
        flight_name: str,
        old_client_code: str,
        new_client_code: str,
    ) -> RenameClientCodeResponse:
        """
        Rename client_code for all records within a specific flight.

        Raises 404 if no records are found for the given flight + old_client_code
        combination — prevents silent no-ops from typos.

        Args:
            session:         Active DB session.
            flight_name:     Target flight scope.
            old_client_code: Current client code to replace.
            new_client_code: Replacement client code.

        Returns:
            RenameClientCodeResponse with the updated row count and new values.

        Raises:
            HTTPException 404: If no records match flight_name + old_client_code.
        """
        updated_count = await ExpectedFlightCargoDAO.rename_client_code(
            session=session,
            flight_name=flight_name,
            old_client_code=old_client_code,
            new_client_code=new_client_code,
        )

        if updated_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=(
                    f"'{old_client_code}' mijoz kodi '{flight_name}' reysida topilmadi."
                ),
            )

        await session.commit()

        return RenameClientCodeResponse(
            updated_count=updated_count,
            flight_name=flight_name,
            old_client_code=old_client_code,
            new_client_code=new_client_code,
        )

    # -------------------------------------------------------------------
    # API 5 — Dynamic delete
    # -------------------------------------------------------------------

    @staticmethod
    async def dynamic_delete(
        session: AsyncSession,
        flight_name: str | None,
        client_code: str | None,
    ) -> DeleteExpectedCargoResponse:
        """
        Delete records matching the provided filter combination.

        See ExpectedFlightCargoDAO.dynamic_delete for the full behaviour matrix.

        Args:
            session:     Active DB session.
            flight_name: Optional flight filter.
            client_code: Optional client filter.

        Returns:
            DeleteExpectedCargoResponse with deleted row count.

        Raises:
            HTTPException 400: If neither filter is supplied.
            HTTPException 404: If the filter matches zero records.
        """
        try:
            deleted_count = await ExpectedFlightCargoDAO.dynamic_delete(
                session=session,
                flight_name=flight_name,
                client_code=client_code,
            )
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(exc),
            ) from exc

        if deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Berilgan filtrlarga mos yozuv topilmadi.",
            )

        await session.commit()

        return DeleteExpectedCargoResponse(deleted_count=deleted_count)

    # -------------------------------------------------------------------
    # API 6 — Excel export
    # -------------------------------------------------------------------

    @staticmethod
    async def generate_excel(
        session: AsyncSession,
        flight_name: str | None,
    ) -> io.BytesIO:
        """
        Build an Excel workbook from expected cargo records and return it as a
        seeked-to-zero BytesIO buffer ready for streaming.

        Each distinct flight name gets its own worksheet so the file can be
        re-imported flight-by-flight without manual splitting.

        Formatting rule: within a client group on a sheet, the first row carries
        client_code; subsequent rows in the same group leave that column blank.

        Args:
            session:     Active DB session.
            flight_name: Optional filter.  None → all flights, each as a sheet.

        Returns:
            BytesIO buffer positioned at byte 0.
        """
        records = await ExpectedFlightCargoDAO.get_all_for_export(
            session=session,
            flight_name=flight_name,
        )

        # Group records by flight_name while preserving insertion order.
        flights: dict[str, list] = {}
        for record in records:
            flights.setdefault(record.flight_name, []).append(record)

        wb = Workbook()
        # Remove the default empty sheet that openpyxl always creates.
        wb.remove(wb.active)  # type: ignore[arg-type]

        for flight, flight_records in flights.items():
            ws = wb.create_sheet(title=_safe_sheet_name(flight))

            # ── Header row ────────────────────────────────────────────
            ws.append(_EXCEL_HEADERS)
            header_row = ws[1]
            for cell in header_row:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Freeze the header row so it stays visible while scrolling.
            ws.freeze_panes = "A2"

            # ── Data rows ─────────────────────────────────────────────
            current_client: str | None = None
            for row_number, record in enumerate(flight_records, start=1):
                is_first_in_client = record.client_code != current_client
                current_client = record.client_code

                # Only the first row of each client group shows the client_code.
                # flight_name column is omitted — it is already the sheet title.
                client_cell = record.client_code if is_first_in_client else ""

                ws.append([
                    row_number,
                    record.flight_name,
                    client_cell,
                    record.track_code,
                    _format_datetime(record.created_at),
                    _format_datetime(record.updated_at),
                ])

                # Lightly shade the first row of each new client group.
                if is_first_in_client:
                    excel_row = ws[ws.max_row]
                    for cell in excel_row:
                        cell.fill = _GROUP_FILL

            # ── Column widths ─────────────────────────────────────────
            column_widths = [6, 20, 15, 30, 20, 20]
            for col_index, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(col_index)].width = width

        # Guard: if there were no records at all, add a placeholder sheet so the
        # workbook is still a valid xlsx file (openpyxl requires at least one sheet).
        if not wb.worksheets:
            wb.create_sheet(title="Ma'lumot yo'q")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # -------------------------------------------------------------------
    # API 7 — Resolve client by track code
    # -------------------------------------------------------------------

    @staticmethod
    async def resolve_client(
        session: AsyncSession,
        track_code: str,
        flight_name: str | None,
    ) -> ResolvedClientResponse:
        """
        Bridge endpoint: look up a track code in cargo_items and return the owning Client.

        Lookup order:
          1. Find CargoItem rows by track_code via CargoItemDAO.get_by_track_code().
          2. If flight_name is provided, narrow to rows that match that flight.
          3. Take the most recent matching row (already ordered by created_at DESC).
          4. Resolve cargo_item.client_id → Client via ClientDAO.get_by_client_code()
             which honours the priority: extra_code > client_code (legacy excluded).

        Args:
            session:     Active DB session.
            track_code:  Scanned tracking code.
            flight_name: Optional flight scope to narrow the lookup.

        Returns:
            ResolvedClientResponse with client_id, codes, name, phone.

        Raises:
            HTTPException 404: If the track code or client is not found.
        """
        cargo_items = await CargoItemDAO.get_by_track_code(
            session=session,
            track_code=track_code,
        )

        # Narrow by flight when the caller supplies one.
        if flight_name:
            normalized_flight = flight_name.strip().lower()
            cargo_items = [
                item for item in cargo_items
                if item.flight_name and item.flight_name.strip().lower() == normalized_flight
            ]

        if not cargo_items:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"'{track_code}' trek kodi yuklar ro'yxatida topilmadi.",
            )

        # Take the most recent record (list is already ordered by created_at DESC).
        cargo_item = cargo_items[0]

        # Guard: if this track_code is already registered in expected_flight_cargos,
        # it has been sent to the backend before — refuse to process it again.
        existing_expected = await ExpectedFlightCargoDAO.find_by_track_code(
            session=session,
            track_code=track_code,
            flight_name=flight_name,
        )
        if existing_expected:
            raise HTTPException(
                status_code=409,
                detail={
                    "detail": f"'{track_code}' trek kodi allaqachon yuborilgan.",
                    "track_code": existing_expected.track_code,
                    "flight_name": existing_expected.flight_name,
                },
            )

        client = await ClientDAO.get_by_client_code(session, cargo_item.client_id)

        if client is None:
            # Client row is absent — the cargo record is still valid, so we
            # return what we know from the cargo table without raising 404.
            logger.info(
                "resolve_client: track_code=%r maps to client_code=%r "
                "which has no registered Client row",
                track_code,
                cargo_item.client_id,
            )
            return ResolvedClientResponse(
                client_id=None,
                client_code=cargo_item.client_id,
                full_name=None,
                phone=None,
                track_code=cargo_item.track_code,
                flight_name=cargo_item.flight_name,
            )

        # Return the raw client_id stored in cargo_items (the code used at scan time).
        # active_code = client.extra_code or client.client_code or cargo_item.client_id

        return ResolvedClientResponse(
            client_id=client.id,
            client_code=cargo_item.client_id,
            full_name=client.full_name,
            phone=client.phone,
            track_code=cargo_item.track_code,
            flight_name=cargo_item.flight_name,
        )

    # -------------------------------------------------------------------
    # API 8 — Summary stats
    # -------------------------------------------------------------------

    @staticmethod
    async def get_summary_stats(session: AsyncSession) -> ExpectedCargoSummaryStats:
        """
        Return aggregate totals for the entire expected cargo table.

        Args:
            session: Active DB session.

        Returns:
            ExpectedCargoSummaryStats with record, flight, and client totals.
        """
        stats = await ExpectedFlightCargoDAO.get_summary_stats(session)
        return ExpectedCargoSummaryStats(
            total_records=stats.total_records,
            total_unique_flights=stats.total_unique_flights,
            total_unique_clients=stats.total_unique_clients,
        )

    # -------------------------------------------------------------------
    # API 9 — Stats by flight (paginated)
    # -------------------------------------------------------------------

    @staticmethod
    async def get_stats_by_flight(
        session: AsyncSession,
        page: int,
        size: int,
        client_code: str | None,
    ) -> PaginatedFlightStatsResponse:
        """
        Return paginated per-flight statistics.

        Rows are ordered by track_code_count descending so the busiest
        flights appear first.  Optionally filtered to flights that contain
        a specific client.

        Args:
            session:     Active DB session.
            page:        1-based page number.
            size:        Rows per page.
            client_code: Optional — restrict to flights containing this client.

        Returns:
            PaginatedFlightStatsResponse.
        """
        rows, total = await ExpectedFlightCargoDAO.get_stats_by_flight(
            session=session,
            page=page,
            size=size,
            client_code=client_code,
        )
        return PaginatedFlightStatsResponse(
            items=[
                FlightStatItem(
                    flight_name=row.flight_name,
                    client_count=row.client_count,
                    track_code_count=row.track_code_count,
                )
                for row in rows
            ],
            total=total,
            page=page,
            size=size,
            total_pages=math.ceil(total / size) if total else 0,
        )

    # -------------------------------------------------------------------
    # API 10 — Stats by client (paginated)
    # -------------------------------------------------------------------

    @staticmethod
    async def get_stats_by_client(
        session: AsyncSession,
        page: int,
        size: int,
        flight_name: str | None,
    ) -> PaginatedClientStatsResponse:
        """
        Return paginated per-client statistics.

        Rows are ordered by track_code_count descending so clients with
        the most cargo appear first.  Optionally filtered to one flight.

        Args:
            session:     Active DB session.
            page:        1-based page number.
            size:        Rows per page.
            flight_name: Optional — restrict to clients in this flight.

        Returns:
            PaginatedClientStatsResponse.
        """
        rows, total = await ExpectedFlightCargoDAO.get_stats_by_client(
            session=session,
            page=page,
            size=size,
            flight_name=flight_name,
        )
        return PaginatedClientStatsResponse(
            items=[
                ClientStatItem(
                    client_code=row.client_code,
                    flight_count=row.flight_count,
                    track_code_count=row.track_code_count,
                )
                for row in rows
            ],
            total=total,
            page=page,
            size=size,
            total_pages=math.ceil(total / size) if total else 0,
        )

    # -------------------------------------------------------------------
    # API 11 — Summary (collapsed client list for a specific flight)
    # -------------------------------------------------------------------

    @staticmethod
    async def get_clients_summary_by_flight(
        session: AsyncSession,
        flight_name: str,
        page: int,
        size: int,
    ) -> PaginatedClientSummaryResponse:
        """
        Return each client's track code count within a specific flight.

        This is the fast "collapsed view" data source: the frontend renders
        a list of clients with a count badge, then expands on demand by
        calling GET / with flight_name + client_code to load the actual codes.

        Args:
            session:     Active DB session.
            flight_name: Required — the flight to scope the query to.
            page:        1-based page number.
            size:        Rows per page.

        Returns:
            PaginatedClientSummaryResponse.
        """
        rows, total = await ExpectedFlightCargoDAO.get_clients_summary_by_flight(
            session=session,
            flight_name=flight_name,
            page=page,
            size=size,
        )
        return PaginatedClientSummaryResponse(
            flight_name=flight_name,
            items=[
                ClientSummaryItem(
                    client_code=row.client_code,
                    total_track_codes=row.track_code_count,
                )
                for row in rows
            ],
            total=total,
            page=page,
            size=size,
            total_pages=math.ceil(total / size) if total else 0,
        )

    # -------------------------------------------------------------------
    # API 12 — Distinct flight list
    # -------------------------------------------------------------------

    @staticmethod
    async def get_flight_list(session: AsyncSession) -> FlightListResponse:
        """
        Return all distinct flight names with aggregate counts.

        Not paginated — flight names are few in practice and the full list
        is needed to populate dropdowns and navigation tabs.

        Args:
            session: Active DB session.

        Returns:
            FlightListResponse with all flights sorted alphabetically.
        """
        rows = await ExpectedFlightCargoDAO.get_distinct_flights(session=session)
        return FlightListResponse(
            items=[
                FlightListItem(
                    flight_name=row.flight_name,
                    client_count=row.client_count,
                    track_code_count=row.track_code_count,
                )
                for row in rows
            ],
            total=len(rows),
        )

    # -------------------------------------------------------------------
    # API 13 — Empty flight registration
    # -------------------------------------------------------------------

    @staticmethod
    async def create_empty_flight(
        session: AsyncSession,
        flight_name: str,
    ) -> "CreateEmptyFlightResponse":
        """Register a placeholder row so the flight becomes visible in listings.

        Idempotent: a second call for the same ``flight_name`` returns
        ``created=False`` without raising.  Used by the web UI to provision
        ``A-`` ostatka flights without having to open Google Sheets.

        Args:
            session:     Active DB session.
            flight_name: Target flight name.

        Returns:
            CreateEmptyFlightResponse with the canonicalised name and a
            boolean flag indicating whether a new row was created.
        """
        from src.api.schemas.expected_cargo import CreateEmptyFlightResponse

        normalised = flight_name.strip()
        created = await ExpectedFlightCargoDAO.create_empty_flight(
            session=session,
            flight_name=normalised,
        )
        await session.commit()
        return CreateEmptyFlightResponse(flight_name=normalised, created=created)


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _format_datetime(dt: datetime | None) -> str:
    """Format a datetime to a human-readable string for Excel cells."""
    return "" if dt is None else dt.strftime("%Y-%m-%d %H:%M")


# Characters that Excel forbids in sheet names.
_SHEET_NAME_FORBIDDEN = str.maketrans({ch: "_" for ch in r"\/*?:[]"})


def _safe_sheet_name(name: str) -> str:
    """
    Sanitise a flight name so it is a valid Excel worksheet title.

    Excel rules: max 31 chars, must not contain \\ / * ? : [ ].
    """
    sanitised = name.translate(_SHEET_NAME_FORBIDDEN).strip()
    return sanitised[:31] if sanitised else "Sheet"
