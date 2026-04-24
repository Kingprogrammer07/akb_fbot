"""Flight Cargo service for business logic."""
import io
from sqlalchemy.ext.asyncio import AsyncSession
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.infrastructure.database.dao.flight_cargo import FlightCargoDAO
from src.infrastructure.database.models.flight_cargo import FlightCargo


class FlightCargoService:
    """Service layer for FlightCargo operations."""

    async def add_cargo(
        self,
        session: AsyncSession,
        flight_name: str,
        client_id: str,
        photo_file_ids: list[str],
        weight_kg: Decimal | None = None,
        price_per_kg: Decimal | None = None,
        comment: str | None = None
    ) -> dict:
        """
        Add a cargo item to a flight.

        Args:
            session: Database session
            flight_name: Flight/reys name
            client_id: Client code
            photo_file_ids: List of Telegram photo file IDs
            weight_kg: Weight in kg
            price_per_kg: Price per kilogram
            comment: Optional comment

        Returns:
            Dict with success status and cargo data or error
        """
        try:
            cargo = await FlightCargoDAO.create(
                session,
                flight_name=flight_name,
                client_id=client_id,
                photo_file_ids=photo_file_ids,
                weight_kg=weight_kg,
                price_per_kg=price_per_kg,
                comment=comment
            )
            await session.commit()

            return {
                'success': True,
                'cargo': cargo
            }
        except Exception as e:
            await session.rollback()
            return {
                'success': False,
                'error': 'creation_failed',
                'message': str(e)
            }

    async def get_cargo_by_id(
        self,
        session: AsyncSession,
        cargo_id: int
    ) -> FlightCargo | None:
        """Get cargo item by ID."""
        return await FlightCargoDAO.get_by_id(session, cargo_id)

    async def get_flight_cargos(
        self,
        session: AsyncSession,
        flight_name: str,
        limit: int = 1000,
        offset: int = 0,
        search: str | None = None,
    ) -> dict:
        """
        Get cargo items for a flight with optional client-ID search.

        The ``search`` filter is forwarded to all three DAO queries so that
        ``total`` and ``unique_clients`` always reflect the filtered result
        set — critical for correct pagination on the frontend.

        Args:
            session:     Async DB session.
            flight_name: Flight name (normalised to upper in the DAO).
            limit:       Page size.
            offset:      Page offset (calculated by the caller as (page-1)*size).
            search:      Optional partial ``client_id`` match (case-insensitive).

        Returns:
            Dict with ``cargos``, ``total``, ``unique_clients``, ``limit``, ``offset``.
        """
        # Sequential queries — AsyncSession is not safe for concurrent use.
        # asyncio.gather on the same session causes IllegalStateChangeError.
        cargos = await FlightCargoDAO.get_by_flight(session, flight_name, limit, offset, search=search)
        total = await FlightCargoDAO.count_by_flight(session, flight_name, search=search)
        unique_clients = await FlightCargoDAO.count_unique_clients_by_flight(session, flight_name, search=search)
        sent_count = await FlightCargoDAO.count_sent_by_flight(session, flight_name)
        unsent_count = await FlightCargoDAO.count_unsent_by_flight(session, flight_name)

        return {
            'cargos': cargos,
            'total': total,
            'unique_clients': unique_clients,
            'sent_count': sent_count,
            'unsent_count': unsent_count,
            'limit': limit,
            'offset': offset
        }

    async def get_client_cargos(
        self,
        session: AsyncSession,
        flight_name: str,
        client_id: str,
        limit: int = 100,
        offset: int = 0
    ) -> dict:
        """
        Get all cargo items for a client in a flight.

        Returns:
            Dict with cargos list and total count
        """
        cargos = await FlightCargoDAO.get_by_client(
            session,
            flight_name,
            client_id,
            limit,
            offset
        )

        return {
            'cargos': cargos,
            'total': len(cargos),
            'limit': limit,
            'offset': offset
        }

    async def get_all_client_cargos(
        self,
        session: AsyncSession,
        client_id: str,
        limit: int = 100,
        offset: int = 0
    ) -> dict:
        """
        Get all cargo items for a client across all flights.

        Returns:
            Dict with cargos list and total count
        """
        cargos = await FlightCargoDAO.get_all_by_client(
            session,
            client_id,
            limit,
            offset
        )

        return {
            'cargos': cargos,
            'total': len(cargos),
            'limit': limit,
            'offset': offset
        }

    async def delete_cargo(
        self,
        session: AsyncSession,
        cargo_id: int
    ) -> dict:
        """
        Delete a cargo item.

        Returns:
            Dict with success status
        """
        try:
            success = await FlightCargoDAO.delete_by_id(session, cargo_id)
            if not success:
                return {
                    'success': False,
                    'error': 'cargo_not_found',
                    'message': 'Cargo item not found'
                }

            await session.commit()
            return {'success': True}
        except Exception as e:
            await session.rollback()
            return {
                'success': False,
                'error': 'deletion_failed',
                'message': str(e)
            }

    async def delete_flight_cargos(
        self,
        session: AsyncSession,
        flight_name: str
    ) -> dict:
        """
        Delete all cargo items for a flight.

        Returns:
            Dict with success status and deleted count
        """
        try:
            deleted_count = await FlightCargoDAO.delete_by_flight(session, flight_name)

            if deleted_count == 0:
                return {
                    'success': False,
                    'error': 'no_cargos_found',
                    'message': f'No cargo items found for flight {flight_name}'
                }

            await session.commit()
            return {
                'success': True,
                'deleted_count': deleted_count
            }
        except Exception as e:
            await session.rollback()
            return {
                'success': False,
                'error': 'deletion_failed',
                'message': str(e)
            }

    async def generate_flight_export(
        self,
        session: AsyncSession,
        flight_name: str
    ) -> io.BytesIO:
        """
        Generate an Excel report for a flight's cargo data.

        Creates a styled workbook in memory with all cargo fields
        except photo_file_ids. Returns a BytesIO buffer.

        Args:
            session: Database session
            flight_name: Flight name

        Returns:
            BytesIO buffer containing the Excel file

        Raises:
            ValueError: If no cargo data found for the flight
        """
        data = await FlightCargoDAO.get_export_data_by_flight(session, flight_name)

        if not data:
            raise ValueError(f"No cargo data found for flight {flight_name}")

        wb = Workbook()
        ws = wb.active
        ws.title = flight_name

        # --- Header row ---
        headers = [
            "ID", "Flight", "Client ID", "Weight (kg)",
            "Price/kg", "Comment", "Sent", "Sent Web", "Sent Web Date",
            "Created At", "Updated At"
        ]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # --- Data rows ---
        date_fmt = "%Y-%m-%d %H:%M"
        for row_idx, row_data in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=row_data["id"])
            ws.cell(row=row_idx, column=2, value=row_data["flight_name"])
            ws.cell(row=row_idx, column=3, value=row_data["client_id"])
            ws.cell(row=row_idx, column=4, value=row_data["weight_kg"])
            ws.cell(row=row_idx, column=5, value=row_data["price_per_kg"])
            ws.cell(row=row_idx, column=6, value=row_data["comment"] or "")
            ws.cell(row=row_idx, column=7, value="Yes" if row_data["is_sent"] else "No")
            ws.cell(row=row_idx, column=8, value="Yes" if row_data["is_sent_web"] else "No")
            ws.cell(
                row=row_idx, column=9,
                value=row_data["is_sent_web_date"].strftime(date_fmt) if row_data["is_sent_web_date"] else ""
            )
            ws.cell(
                row=row_idx, column=10,
                value=row_data["created_at"].strftime(date_fmt) if row_data["created_at"] else ""
            )
            ws.cell(
                row=row_idx, column=11,
                value=row_data["updated_at"].strftime(date_fmt) if row_data["updated_at"] else ""
            )

        # --- Auto-width columns ---
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

        # --- Save to buffer ---
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
